from __future__ import annotations

import csv
import io
from io import BytesIO
import json
import re
import time
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any

import requests
from bs4 import BeautifulSoup, NavigableString
from pypdf import PdfReader
from openpyxl import load_workbook

DATA_FILE = Path("data/uk_macro.json")
DEBUG_DIR = Path("debug/uk_macro_sources")
USER_AGENT = "Mozilla/5.0 (compatible; UKMacroDashboard/1.0)"

SESSION = requests.Session()
SESSION.headers.update({
    "User-Agent": USER_AGENT,
    "Accept-Language": "en-GB,en;q=0.9",
})

ONS = {
    "ukhca9iq": ("MM23", "DKO8", "economy/inflationandpriceindices", 0),
    "ukhpsery": ("MM23", "D7NN", "economy/inflationandpriceindices", 0),
    "ukueilor": ("LMS", "MGSX", "employmentandlabourmarket/peoplenotinwork/unemployment", 1),
    "ukuer": ("UNEM", "BCJE", "employmentandlabourmarket/peoplenotinwork/outofworkbenefits", 0),
    "ukawmwho": ("LMS", "KAC3", "employmentandlabourmarket/peopleinwork/earningsandworkinghours", 0),
    "ukawxprm": ("LMS", "KAJ4", "employmentandlabourmarket/peopleinwork/earningsandworkinghours", 0),
    # AP2Y Raw月份是3個月統計期間的中間月；Dashboard採發布月份，所以+2個月。
    "ukvaap2y": ("UNEM", "AP2Y", "employmentandlabourmarket/peopleinwork/employmentandemployeetypes", 2),
    "uklfjpc5": ("UNEM", "JPC5", "employmentandlabourmarket/peoplenotinwork/unemployment", 1),
    "ukgdm3m": ("MGDP", "ECYX", "economy/grossdomesticproductgdp", 0),
}

LEVELS = {
    "ukgrabiy": ("QNA", "ABMI", "economy/grossdomesticproductgdp"),
    "ukgeabry": ("PN2", "ABJR", "economy/nationalaccounts/satelliteaccounts"),
    "ukgvnpqy": ("UKEA", "NPQT", "economy/grossdomesticproductgdp"),
}

MONTHS = {
    name: index
    for index, name in enumerate(
        [
            "january", "february", "march", "april", "may", "june",
            "july", "august", "september", "october", "november", "december",
        ],
        1,
    )
}
MONTH_ABBR = {name[:3]: number for name, number in MONTHS.items()}

RETAIL_XLSX_URL = (
    "https://www.ons.gov.uk/file?uri=/businessindustryandtrade/retailindustry/"
    "datasets/retailsalesindexreferencetables/current/mainreferencetables.xlsx"
)
GFK_URL = "https://tradingeconomics.com/united-kingdom/consumer-confidence"
MANUFACTURING_PMI_URL = (
    "https://www.investing.com/economic-calendar/"
    "united-kingdom-manufacturing-purchasing-managers-index-(pmi)-204"
)
SERVICES_PMI_URL = (
    "https://www.investing.com/economic-calendar/"
    "united-kingdom-services-purchasing-managers-index-(pmi)-274"
)


def get(url: str, **kwargs: Any) -> requests.Response:
    last_response: requests.Response | None = None
    timeout = kwargs.pop("timeout", 25)

    for attempt in range(3):
        print(
            f"[HTTP] attempt {attempt + 1}/3: {url}",
            flush=True,
        )
        try:
            response = SESSION.get(
                url,
                timeout=timeout,
                **kwargs,
            )
        except requests.RequestException as error:
            print(
                f"[HTTP] request error: {type(error).__name__}: {error}",
                flush=True,
            )
            if attempt == 2:
                raise
            time.sleep(2 * (2**attempt))
            continue

        last_response = response
        print(
            f"[HTTP] status={response.status_code} "
            f"bytes={len(response.content)}",
            flush=True,
        )

        if response.status_code < 400:
            return response

        if response.status_code not in {429, 500, 502, 503, 504}:
            response.raise_for_status()

        if attempt < 2:
            time.sleep(2 * (2**attempt))

    assert last_response is not None
    last_response.raise_for_status()
    return last_response


def save_debug(name: str, payload: Any) -> None:
    DEBUG_DIR.mkdir(parents=True, exist_ok=True)
    path = DEBUG_DIR / name
    if isinstance(payload, str):
        path.write_text(payload, encoding="utf-8")
    else:
        path.write_text(
            json.dumps(payload, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )


def debug_print(label: str, payload: Any) -> None:
    print(f"\n[DEBUG {label}]", flush=True)
    print(
        json.dumps(payload, ensure_ascii=False, indent=2),
        flush=True,
    )


def period(value: str) -> str | None:
    value = " ".join(str(value).upper().split())
    aliases = {
        "JAN": "01", "FEB": "02", "MAR": "03", "APR": "04",
        "MAY": "05", "JUN": "06", "JUL": "07", "AUG": "08",
        "SEP": "09", "OCT": "10", "NOV": "11", "DEC": "12",
    }
    month_match = re.fullmatch(r"(\d{4}) (" + "|".join(aliases) + r")", value)
    if month_match:
        return f"{month_match[1]}-{aliases[month_match[2]]}-01"
    quarter_match = re.fullmatch(r"(\d{4}) Q([1-4])", value)
    if quarter_match:
        return f"{quarter_match[1]}-{int(quarter_match[2]) * 3:02d}-01"
    return None


def shift_month(date_value: str, offset: int = 1) -> str:
    year, month = map(int, date_value[:7].split("-"))
    serial = year * 12 + month - 1 + offset
    return f"{serial // 12:04d}-{serial % 12 + 1:02d}-01"


def ons_series(dataset: str, cdid: str, path: str) -> list[dict[str, Any]]:
    edition = {"ABJR": "pn2", "ABMI": "qna", "NPQT": "ukea"}.get(
        cdid,
        dataset.lower(),
    )
    url = (
        "https://www.ons.gov.uk/generator?format=csv&uri=/"
        f"{path}/timeseries/{cdid.lower()}/{edition}"
    )
    raw = get(url).content.decode("utf-8-sig", "replace")
    rows = csv.reader(io.StringIO(raw))
    output = []
    for row in rows:
        if len(row) < 2:
            continue
        date_value = period(row[0])
        raw_value = row[1].replace(",", "").strip()
        if date_value and re.fullmatch(r"-?\d+(?:\.\d+)?", raw_value):
            output.append({
                "date": date_value,
                "value": float(raw_value),
                "source_url": url,
            })
    if not output:
        raise RuntimeError(f"ONS {cdid} returned no data")

    # ONS rate-limits rapid consecutive Generator requests.
    time.sleep(1.5)
    return output


def year_over_year(levels: list[dict[str, Any]]) -> list[dict[str, Any]]:
    indexed = {point["date"]: point for point in levels}
    output = []
    for date_value, point in indexed.items():
        previous_date = f"{int(date_value[:4]) - 1}{date_value[4:]}"
        previous = indexed.get(previous_date)
        if previous and previous["value"]:
            output.append({
                "date": date_value,
                "value": (point["value"] / previous["value"] - 1) * 100,
                "source_url": point["source_url"],
            })
    return output


def by_id(database: dict[str, Any], series_id: str) -> dict[str, Any] | None:
    return next(
        (item for item in database["series"] if item["id"] == series_id),
        None,
    )


def month_key(value: Any) -> str:
    match = re.match(r"^(\d{4})-(\d{2})", str(value or "").strip())
    return f"{match.group(1)}-{match.group(2)}" if match else ""


def merge(
    database: dict[str, Any],
    series_id: str,
    points: list[dict[str, Any]],
    release_type: str | None = None,
    replace_source_range: bool = False,
) -> tuple[int, int]:
    series = by_id(database, series_id)
    if not series:
        raise KeyError(series_id)

    old = {
        month_key(point.get("date")): {
            **point,
            "date": month_key(point.get("date")) + "-01",
        }
        for point in series.get("data", [])
        if month_key(point.get("date"))
    }

    normalized_points: dict[str, dict[str, Any]] = {}
    for point in points:
        key = month_key(point.get("date"))
        if not key:
            continue
        candidate = {**point, "date": key + "-01"}
        if release_type:
            candidate["release_type"] = release_type
        normalized_points[key] = candidate

    if not normalized_points:
        return 0, 0

    added = revised = 0

    if replace_source_range:
        # Official ONS sources are authoritative over every overlapping period.
        # Preserve user history before the first date already stored in JSON,
        # overwrite all overlapping dates, and append newer official periods.
        # This prevents an ONS series beginning in 1985 from unexpectedly
        # extending a user file whose intended history begins in 2015.
        earliest_existing = min(old) if old else min(normalized_points)
        keys_to_apply = [
            key for key in normalized_points
            if key >= earliest_existing
        ]
    else:
        # Non-official/scraped sources only add new periods or revise the latest
        # available period, avoiding backfilling unrelated old history.
        if old:
            latest_existing = max(old)
            keys_to_apply = [
                key for key in normalized_points
                if key >= latest_existing
            ]
        else:
            keys_to_apply = list(normalized_points)

    for key in sorted(keys_to_apply):
        candidate = normalized_points[key]
        current = old.get(key)

        if current is None:
            old[key] = candidate
            added += 1
            continue

        if (
            current.get("release_type") == "final"
            and candidate.get("release_type") == "flash"
        ):
            continue

        changed = (
            current.get("value") != candidate.get("value")
            or (
                candidate.get("release_type") == "final"
                and current.get("release_type") != "final"
            )
        )
        if changed:
            old[key] = {**current, **candidate}
            revised += 1
        elif replace_source_range:
            # Refresh source metadata even when the official value is unchanged.
            old[key] = {**current, **candidate}

    series["data"] = sorted(old.values(), key=lambda item: item["date"])
    return added, revised


def clean_cell(value: str) -> str:
    return re.sub(r"\s+", " ", value).strip()


def extract_tables(html: str) -> list[list[list[str]]]:
    soup = BeautifulSoup(html, "html.parser")
    tables = []
    for table in soup.find_all("table"):
        rows = []
        for row in table.find_all("tr"):
            cells = [
                clean_cell(cell.get_text(" ", strip=True))
                for cell in row.find_all(["th", "td"])
            ]
            if cells:
                rows.append(cells)
        if rows:
            tables.append(rows)
    return tables


def number_or_none(value: str) -> float | None:
    cleaned = value.replace(",", "").replace("%", "").strip()
    if not cleaned or cleaned in {"-", "N/A", "n/a"}:
        return None
    match = re.fullmatch(r"[+-]?\d+(?:\.\d+)?", cleaned)
    return float(cleaned) if match else None


def te_reference_month(release_date: str, reference: str) -> str | None:
    date_match = re.match(r"(20\d{2})-(\d{2})-(\d{2})", release_date)
    if not date_match:
        return None
    reference_no = MONTH_ABBR.get(reference.strip().lower()[:3])
    if not reference_no:
        return None
    year = int(date_match.group(1))
    release_month = int(date_match.group(2))
    if reference_no > release_month + 1:
        year -= 1
    return f"{year:04d}-{reference_no:02d}-01"


def update_te_table(
    database: dict[str, Any],
    series_id: str,
    url: str,
    indicator_name: str,
    debug_name: str,
) -> tuple[int, int]:
    response = get(url)
    tables = extract_tables(response.text)
    matched_rows = []
    points = []

    for table_no, table in enumerate(tables, 1):
        for row_no, cells in enumerate(table, 1):
            if len(cells) < 6 or indicator_name.lower() not in cells[2].lower():
                continue
            release_date, release_time, indicator, reference, actual = cells[:5]
            matched_rows.append({
                "table_no": table_no,
                "row_no": row_no,
                "cells": cells,
                "release_date": release_date,
                "reference": reference,
                "actual_raw": actual,
            })
            actual_value = number_or_none(actual)
            date_value = te_reference_month(release_date, reference)
            if actual_value is None or not date_value:
                continue
            points.append({
                "date": date_value,
                "value": actual_value,
                "source_url": url,
            })

    save_debug(f"{debug_name}_raw.html", response.text)
    save_debug(f"{debug_name}_tables.json", tables)
    save_debug(f"{debug_name}_matched_rows.json", matched_rows)
    debug_print(debug_name, matched_rows)

    if not points:
        raise RuntimeError(f"No actual rows found for {indicator_name}")
    return merge(database, series_id, points)


def normalize_excel_label(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", " ", str(value or "").lower()).strip()


def excel_month(value: Any) -> str | None:
    if isinstance(value, datetime):
        return f"{value.year:04d}-{value.month:02d}-01"

    text = clean_cell(str(value or ""))
    if not text:
        return None

    patterns = [
        r"^(20\d{2})[-/ ](0?[1-9]|1[0-2])(?:[-/ ]\d{1,2})?$",
        r"^(20\d{2})\s*M(0?[1-9]|1[0-2])$",
    ]
    for pattern_value in patterns:
        match = re.fullmatch(pattern_value, text, re.I)
        if match:
            return f"{int(match.group(1)):04d}-{int(match.group(2)):02d}-01"

    month_year = re.fullmatch(
        r"(January|February|March|April|May|June|July|August|September|October|November|December)"
        r"\s+(20\d{2})",
        text,
        re.I,
    )
    if month_year:
        month = MONTHS[month_year.group(1).lower()]
        return f"{int(month_year.group(2)):04d}-{month:02d}-01"

    year_month = re.fullmatch(
        r"(20\d{2})\s+"
        r"(January|February|March|April|May|June|July|August|September|October|November|December|"
        r"Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)",
        text,
        re.I,
    )
    if year_month:
        month = MONTH_ABBR[year_month.group(2).lower()[:3]]
        return f"{int(year_month.group(1)):04d}-{month:02d}-01"

    return None


def discover_kpsa1_sheet(workbook: Any) -> Any:
    for sheet_name in workbook.sheetnames:
        normalized = re.sub(r"[^a-z0-9]", "", sheet_name.lower())
        if normalized == "kpsa1":
            return workbook[sheet_name]
    raise RuntimeError(
        "ONS retail workbook does not contain the KPSA 1 worksheet; "
        f"available sheets: {workbook.sheetnames}"
    )


def extract_kpsa1_retail_ex_fuel(
    sheet: Any,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """Extract KPSA 1 year-on-year retail sales excluding automotive fuel.

    KPSA 1 contains more than one table with the same category headings.  The
    required series is specifically the upper table headed
    "Percentage change on same month a year earlier" and dataset code J45U.
    The lower "Revision to index numbers" table must never be treated as the
    level of the year-on-year series.
    """
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise RuntimeError("KPSA 1 worksheet is empty")

    wanted_section = "percentage change on same month a year earlier"
    stop_section = "revision to index numbers"
    wanted_label = "all retailing excluding automotive fuel"
    wanted_code = "j45u"

    section_start = None
    section_end = len(rows)
    for row_index, row in enumerate(rows):
        row_text = " | ".join(normalize_excel_label(value) for value in row if value is not None)
        if section_start is None and wanted_section in row_text:
            section_start = row_index
            continue
        if section_start is not None and stop_section in row_text:
            section_end = row_index
            break

    if section_start is None:
        raise RuntimeError(
            "KPSA 1 does not contain the section "
            "'Percentage change on same month a year earlier'"
        )

    # Identify the required column inside the selected upper block. Prefer the
    # stable ONS dataset identifier J45U; verify that the same column carries
    # the expected category label.
    target_column = None
    code_row = None
    label_matches: set[int] = set()
    for row_index in range(section_start, section_end):
        row = rows[row_index]
        for column_index, value in enumerate(row):
            normalized = normalize_excel_label(value)
            if wanted_label in normalized:
                label_matches.add(column_index)
            if normalized == wanted_code:
                target_column = column_index
                code_row = row_index
                break
        if target_column is not None:
            break

    if target_column is None:
        if len(label_matches) == 1:
            target_column = next(iter(label_matches))
        else:
            raise RuntimeError(
                "Could not uniquely locate J45U / All retailing excluding "
                "automotive fuel in the required KPSA 1 section"
            )

    if label_matches and target_column not in label_matches:
        raise RuntimeError(
            "KPSA 1 J45U column does not align with the expected "
            "All retailing excluding automotive fuel heading"
        )

    candidates: list[dict[str, Any]] = []
    for row_index in range((code_row + 1) if code_row is not None else section_start + 1, section_end):
        row = rows[row_index]
        if target_column >= len(row):
            continue

        date_value = None
        date_column = None
        # The time period is normally in the first column, but scan only the
        # columns to the left of J45U to tolerate minor ONS layout changes.
        for column_index in range(0, min(target_column, len(row))):
            parsed_date = excel_month(row[column_index])
            if parsed_date:
                date_value = parsed_date
                date_column = column_index
                break

        numeric_value = number_or_none(str(row[target_column] or ""))
        if date_value and numeric_value is not None:
            candidates.append({
                "section": "Percentage change on same month a year earlier",
                "dataset_code": "J45U",
                "orientation": "dates_down_rows",
                "date_cell": f"R{row_index + 1}C{date_column + 1}",
                "value_cell": f"R{row_index + 1}C{target_column + 1}",
                "date": date_value,
                "value": numeric_value,
            })

    if not candidates:
        raise RuntimeError(
            "Found the required KPSA 1 upper section and J45U column, "
            "but no dated numeric observations were extracted"
        )

    by_month: dict[str, dict[str, Any]] = {}
    for row in candidates:
        by_month[row["date"]] = {
            "date": row["date"],
            "value": row["value"],
            "source_url": RETAIL_XLSX_URL,
            "source_sheet": sheet.title,
            "source_cell": row["value_cell"],
            "source_section": row["section"],
            "dataset_code": row["dataset_code"],
            "measure": (
                "KPSA 1 - Percentage change on same month a year earlier; "
                "All retailing excluding automotive fuel (J45U)"
            ),
        }

    points = sorted(by_month.values(), key=lambda item: item["date"])
    return points, candidates

def update_retail(database: dict[str, Any]) -> tuple[int, int]:
    response = get(RETAIL_XLSX_URL)
    if not response.content.startswith(b"PK"):
        raise RuntimeError(
            "ONS retail download did not return an XLSX file "
            f"(content-type={response.headers.get('content-type')!r}, "
            f"bytes={len(response.content)})"
        )

    workbook = load_workbook(
        BytesIO(response.content),
        read_only=True,
        data_only=True,
    )
    sheet = discover_kpsa1_sheet(workbook)
    points, candidates = extract_kpsa1_retail_ex_fuel(sheet)

    debug_payload = {
        "source_url": RETAIL_XLSX_URL,
        "content_type": response.headers.get("content-type"),
        "content_bytes": len(response.content),
        "sheet_names": workbook.sheetnames,
        "selected_sheet": sheet.title,
        "candidate_count": len(candidates),
        "candidate_preview": candidates[-24:],
        "observation_count": len(points),
        "latest_observations": points[-24:],
    }
    save_debug("retail_ons_kpsa1_debug.json", debug_payload)
    debug_print("retail_ons_kpsa1", debug_payload)

    return merge(
        database, "ukrvayoy", points, replace_source_range=True
    )


def update_gfk(database: dict[str, Any]) -> tuple[int, int]:
    return update_te_table(
        database,
        "ukcci",
        GFK_URL,
        "GfK Consumer Confidence",
        "gfk_consumer_confidence",
    )


SP_RELEASES_URL = "https://www.pmi.spglobal.com/Public/Release/PressReleases"


def response_to_text(response: requests.Response) -> str:
    content_type = (response.headers.get("content-type") or "").lower()
    if response.content.startswith(b"%PDF") or "application/pdf" in content_type:
        reader = PdfReader(BytesIO(response.content))
        return "\n".join(page.extract_text() or "" for page in reader.pages)
    return BeautifulSoup(response.text, "html.parser").get_text("\n", strip=True)


def preceding_release_context(anchor: Any) -> str:
    parts = []
    for element in anchor.previous_elements:
        if isinstance(element, NavigableString):
            text = clean_cell(str(element))
            if text:
                parts.append(text)
        if len(" ".join(parts)) >= 280:
            break
    return " ".join(reversed(parts[-20:]))


def parse_sp_release_date(value: str) -> datetime | None:
    text = clean_cell(value)
    if not text:
        return None
    for format_value in ("%B %d %Y", "%b %d %Y"):
        try:
            return datetime.strptime(text, format_value).replace(
                tzinfo=timezone.utc,
            )
        except ValueError:
            continue
    return None


def discover_sp_pmi_releases() -> list[dict[str, str]]:
    response = get(SP_RELEASES_URL)
    soup = BeautifulSoup(response.text, "html.parser")
    candidates = []
    seen = set()
    for anchor in soup.find_all("a", href=True):
        href = anchor.get("href", "")
        if "/Public/Home/PressRelease/" not in href:
            continue
        url = requests.compat.urljoin(SP_RELEASES_URL, href)
        if url in seen:
            continue

        # S&P sometimes places the release title inside the link itself and
        # sometimes in the surrounding card. Read both, rather than only the
        # text before the anchor; otherwise the latest Flash PMI can be missed.
        anchor_text = clean_cell(anchor.get_text(" ", strip=True))
        parent_text = clean_cell(
            anchor.parent.get_text(" ", strip=True) if anchor.parent else ""
        )
        previous_text = preceding_release_context(anchor)
        context = clean_cell(" ".join(
            part for part in (anchor_text, parent_text, previous_text) if part
        ))

        title_match = re.search(
            r"(S&P Global\s+(?:Flash\s+)?UK(?:\s+(?:Manufacturing|Services))?\s+PMI(?:[^|]{0,80})?)",
            context,
            re.I,
        )
        if not title_match:
            # Keep a broader fallback for minor S&P title-format changes.
            title_match = re.search(
                r"((?:Flash\s+)?UK(?:\s+(?:Manufacturing|Services))?\s+PMI(?:[^|]{0,80})?)",
                context,
                re.I,
            )
        if not title_match:
            continue
        title = clean_cell(title_match.group(1))

        release_date_match = re.search(
            r"([A-Za-z]+\s+\d{1,2},?\s+20\d{2})\s+\d{2}:\d{2}\s+UTC",
            context,
            re.I,
        )
        release_date = release_date_match.group(1).replace(",", "") if release_date_match else ""
        candidates.append({
            "title": title,
            "url": url,
            "release_date": release_date,
            "index_context": context,
        })
        seen.add(url)

    cutoff = datetime.now(timezone.utc) - timedelta(days=150)
    recent_candidates = []
    for candidate in candidates:
        release_date = parse_sp_release_date(candidate.get("release_date", ""))
        if release_date is not None and release_date >= cutoff:
            recent_candidates.append(candidate)

    # If S&P changes its date text format, process a bounded set of the newest
    # visible UK releases instead of silently returning nothing.
    if not recent_candidates:
        recent_candidates = candidates[:30]
    recent_candidates.sort(
        key=lambda item: parse_sp_release_date(item.get("release_date", ""))
        or datetime.min.replace(tzinfo=timezone.utc),
        reverse=True,
    )
    print(
        f"[S&P PMI] discovered={len(candidates)} "
        f"recent_to_process={len(recent_candidates)}",
        flush=True,
    )
    save_debug("sp_global_release_candidates.json", recent_candidates)
    debug_print("sp_global_release_candidates", recent_candidates)
    return recent_candidates

def extract_reference_month(text: str) -> str | None:
    head = clean_cell(text[:5000])
    matches = list(re.finditer(
        r"\b(January|February|March|April|May|June|July|August|September|October|November|December)\s+(20\d{2})\b",
        head,
        re.I,
    ))
    if not matches:
        return None
    # The report heading normally contains the reference month near the start.
    match = matches[0]
    month = MONTHS[match.group(1).lower()]
    return f"{int(match.group(2)):04d}-{month:02d}"


def extract_pmi_value(text: str, sector: str, release_type: str) -> float | None:
    compact = clean_cell(text)

    if sector == "manufacturing":
        labels = [
            r"S&P Global UK Manufacturing PMI",
            r"UK Manufacturing PMI",
            r"Manufacturing PMI",
            r"Manufacturing Purchasing Managers(?:’|') Index",
        ]
    else:
        labels = [
            r"S&P Global UK Services PMI Business Activity Index",
            r"UK Services PMI Business Activity Index",
            r"Services PMI Business Activity Index",
            r"Services Business Activity Index",
            r"UK Services PMI",
            r"Services PMI",
        ]

    verbs = r"(?:at|posted|registered|rose to|fell to|increased to|decreased to|unchanged at|=|:)"
    patterns = []
    for label in labels:
        patterns.extend([
            rf"{label}[^0-9]{{0,100}}{verbs}\s*([0-9]{{2}}(?:\.[0-9]+)?)",
            rf"{label}[^0-9]{{0,50}}([0-9]{{2}}(?:\.[0-9]+)?)",
        ])

    if release_type == "flash":
        flash_label = (
            r"Flash UK Manufacturing PMI"
            if sector == "manufacturing"
            else r"Flash UK Services PMI"
        )
        patterns.insert(0, rf"{flash_label}[^0-9]{{0,100}}([0-9]{{2}}(?:\.[0-9]+)?)")

    for pattern_value in patterns:
        match = re.search(pattern_value, compact, re.I)
        if match:
            value = float(match.group(1))
            if 20.0 <= value <= 80.0:
                return value
    return None


def month_range(start_month: str, end_month: str) -> list[str]:
    if not start_month or not end_month or start_month > end_month:
        return []
    output = []
    current = start_month + "-01"
    while month_key(current) <= end_month:
        output.append(month_key(current))
        current = shift_month(current, 1)
    return output


def update_sp_global_pmi(database: dict[str, Any]) -> dict[str, tuple[int, int]]:
    candidates = discover_sp_pmi_releases()
    mapping = {
        "manufacturing": "mpmigbma",
        "services": "mpmigbsa",
    }

    # The current calendar month is the latest month that may already have a
    # Flash PMI release. Missing months are attempted one by one; a month with
    # no published/parsable release is logged and skipped without failing the run.
    now_month = datetime.now(timezone.utc).strftime("%Y-%m")
    existing: dict[str, dict[str, dict[str, Any]]] = {}
    needed: dict[str, set[str]] = {}
    ordered_needed: dict[str, list[str]] = {}

    for sector, series_id in mapping.items():
        series = by_id(database, series_id) or {}
        rows = {
            month_key(row.get("date")): row
            for row in series.get("data", [])
            if month_key(row.get("date"))
        }
        existing[sector] = rows

        months_to_try: list[str] = []
        if rows:
            latest_month = max(rows)
            first_missing = month_key(shift_month(latest_month + "-01", 1))
            months_to_try.extend(month_range(first_missing, now_month))
            # Revisit Flash observations because a later final release may now exist.
            months_to_try.extend(
                month for month, row in rows.items()
                if row.get("release_type") == "flash" and month <= now_month
            )
        else:
            months_to_try.append(now_month)

        ordered_needed[sector] = sorted(set(months_to_try))
        needed[sector] = set(ordered_needed[sector])
        latest_display = max(rows) if rows else "none"
        print(
            f"[S&P PMI] {sector}: latest_existing={latest_display}; "
            f"trying_through={now_month}; months={ordered_needed[sector]}",
            flush=True,
        )

    observations = {"manufacturing": [], "services": []}
    release_debug = []
    resolved: set[tuple[str, str]] = set()

    for candidate in candidates:
        title_lower = candidate["title"].lower()
        release_type = "flash" if "flash" in title_lower else "final"
        if "manufacturing" in title_lower:
            candidate_sectors = ["manufacturing"]
        elif "services" in title_lower:
            candidate_sectors = ["services"]
        elif "flash" in title_lower and "uk" in title_lower and "pmi" in title_lower:
            candidate_sectors = ["manufacturing", "services"]
        else:
            continue

        release_dt = parse_sp_release_date(candidate.get("release_date", ""))
        if release_dt is None:
            release_debug.append({
                **candidate,
                "skipped": "release date could not be parsed",
            })
            continue

        if release_type == "flash":
            expected_month = f"{release_dt.year:04d}-{release_dt.month:02d}"
        else:
            expected_month = month_key(shift_month(
                f"{release_dt.year:04d}-{release_dt.month:02d}-01", -1
            ))

        sectors_to_fetch = [
            sector for sector in candidate_sectors
            if expected_month in needed[sector]
            and (sector, expected_month) not in resolved
            and not (
                release_type == "flash"
                and existing[sector].get(expected_month, {}).get("release_type") == "final"
            )
        ]
        if not sectors_to_fetch:
            continue

        print(
            f"[S&P PMI] attempting {expected_month} {release_type}: "
            f"{candidate['title']} -> {sectors_to_fetch}",
            flush=True,
        )
        try:
            response = get(candidate["url"])
            text_value = response_to_text(response)
        except Exception as error:
            release_debug.append({**candidate, "error": str(error)})
            print(
                f"[S&P PMI] skipped {expected_month}; download failed: {error}",
                flush=True,
            )
            continue

        parsed = {
            **candidate,
            "release_type": release_type,
            "reference_month": expected_month,
            "content_type": response.headers.get("content-type"),
            "content_bytes": len(response.content),
            "values": {},
        }
        safe_name = candidate["url"].rstrip("/").split("/")[-1]
        save_debug(f"sp_release_{safe_name}.txt", text_value)

        for sector in sectors_to_fetch:
            value = extract_pmi_value(text_value, sector, release_type)
            parsed["values"][sector] = value
            if value is None:
                print(
                    f"[S&P PMI] {expected_month} {sector}: release found "
                    "but value could not be parsed; skipping",
                    flush=True,
                )
                continue
            observations[sector].append({
                "date": expected_month + "-01",
                "value": value,
                "release_type": release_type,
                "source_url": candidate["url"],
            })
            resolved.add((sector, expected_month))
            print(
                f"[S&P PMI] resolved {expected_month} {sector}={value} "
                f"({release_type})",
                flush=True,
            )
        release_debug.append(parsed)

    selected: dict[str, list[dict[str, Any]]] = {}
    for sector, rows in observations.items():
        best: dict[str, dict[str, Any]] = {}
        for row in rows:
            month = month_key(row["date"])
            current = best.get(month)
            if current is None or (
                row["release_type"] == "final"
                and current.get("release_type") != "final"
            ):
                best[month] = row
        selected[sector] = sorted(best.values(), key=lambda row: row["date"])

    save_debug("sp_global_release_parsed.json", release_debug)
    save_debug("sp_global_pmi_selected.json", selected)
    debug_print("sp_global_release_parsed", release_debug)
    debug_print("sp_global_pmi_selected", selected)

    results = {}
    for sector, series_id in mapping.items():
        added = revised = 0
        selected_months = set()
        for row in selected[sector]:
            selected_months.add(month_key(row["date"]))
            release_type = row["release_type"]
            point = {k: v for k, v in row.items() if k != "release_type"}
            a, r = merge(database, series_id, [point], release_type)
            added += a
            revised += r

        unresolved = [
            month for month in ordered_needed[sector]
            if month not in selected_months
        ]
        for month in unresolved:
            print(
                f"[S&P PMI] {sector} {month}: no available release/value; skipped",
                flush=True,
            )
        results[series_id] = (added, revised)
    return results

def dmp_page_candidates() -> list[tuple[str, str]]:
    now = datetime.now(timezone.utc)
    output = []
    for offset in range(0, 6):
        serial = now.year * 12 + now.month - 1 - offset
        year = serial // 12
        month = serial % 12 + 1
        month_name = list(MONTHS)[month - 1]
        output.append((
            f"{year:04d}-{month:02d}",
            f"https://www.bankofengland.co.uk/decision-maker-panel/{year}/{month_name}-{year}",
        ))
    return output


def parse_dmp_month(value: Any) -> str | None:
    """Parse the date formats used by the BoE DMP workbook."""
    parsed = excel_month(value)
    if parsed:
        return parsed
    if isinstance(value, datetime):
        return f"{value.year:04d}-{value.month:02d}-01"
    text = clean_cell(str(value or ""))
    for fmt in ("%b-%y", "%b %y", "%B-%y", "%B %y", "%Y-%m", "%m/%Y"):
        try:
            parsed_dt = datetime.strptime(text, fmt)
            return f"{parsed_dt.year:04d}-{parsed_dt.month:02d}-01"
        except ValueError:
            pass
    return None


def dmp_excel_candidates() -> list[tuple[str, str, str]]:
    """Build direct official monthly-XLSX URLs from current month backwards."""
    output = []
    for reference_month, page_url in dmp_page_candidates():
        year, month = map(int, reference_month.split("-"))
        month_name = list(MONTHS)[month - 1]
        xlsx_url = (
            "https://www.bankofengland.co.uk/-/media/boe/files/"
            f"decision-maker-panel-survey/{year}/"
            f"monthly-dmp-data-{month_name}-{year}.xlsx"
        )
        output.append((reference_month, page_url, xlsx_url))
    return output


def download_latest_dmp_excel() -> tuple[str, str, str, requests.Response]:
    """Try official XLSX URLs directly; no dependency on parsing the release page."""
    attempts = []
    for reference_month, page_url, xlsx_url in dmp_excel_candidates():
        try:
            response = get(
                xlsx_url,
                headers={
                    "Referer": page_url,
                    "Accept": (
                        "application/vnd.openxmlformats-officedocument."
                        "spreadsheetml.sheet,application/octet-stream,*/*"
                    ),
                },
            )
            is_xlsx = response.content.startswith(b"PK")
            attempts.append({
                "reference_month": reference_month,
                "page_url": page_url,
                "xlsx_url": xlsx_url,
                "status": response.status_code,
                "bytes": len(response.content),
                "content_type": response.headers.get("content-type"),
                "is_xlsx": is_xlsx,
            })
            if is_xlsx:
                save_debug("boe_dmp_excel_download_attempts.json", attempts)
                debug_print("boe_dmp_excel_download_attempts", attempts)
                return reference_month, page_url, xlsx_url, response
        except Exception as error:
            attempts.append({
                "reference_month": reference_month,
                "page_url": page_url,
                "xlsx_url": xlsx_url,
                "error": str(error),
            })
    save_debug("boe_dmp_excel_download_attempts.json", attempts)
    debug_print("boe_dmp_excel_download_attempts", attempts)
    raise RuntimeError("No recent BoE monthly DMP XLSX could be downloaded")


def extract_dmp_single_month_cpi(
    workbook: Any,
    source_url: str,
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    sheet = next(
        (workbook[name] for name in workbook.sheetnames
         if "cpi" in normalize_excel_label(name)
         and "expect" in normalize_excel_label(name)),
        None,
    )
    if sheet is None:
        raise RuntimeError(
            "DMP workbook does not contain a CPI expectations worksheet; "
            f"available sheets: {workbook.sheetnames}"
        )
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise RuntimeError("DMP CPI expectations worksheet is empty")

    # DMP sheets can be laid out either with dates down rows or across columns.
    # Locate every date cell first, then locate the exact metric label using the
    # surrounding row/column text. The requested series is one-year-ahead CPI
    # expectations, single month (never the 3-month average).
    date_cells = []
    for r, row in enumerate(rows):
        for c, value in enumerate(row):
            date_value = parse_dmp_month(value)
            if date_value:
                date_cells.append((r, c, date_value))
    if not date_cells:
        raise RuntimeError("No monthly dates found in DMP CPI expectations sheet")

    def metric_score(text: str) -> int:
        label = normalize_excel_label(text)
        score = 0
        if re.search(r"(?:1|one) year ahead", label):
            score += 8
        if "expect" in label:
            score += 4
        if "single month" in label:
            score += 8
        if "3 month" in label or "three month" in label:
            score -= 20
        if "current cpi" in label or re.search(r"3 year ahead", label):
            score -= 20
        return score

    candidates = []
    max_columns = max(len(row) for row in rows)
    # Row-oriented metric: metric label on a row, dates across columns.
    for r, row in enumerate(rows):
        row_label = " ".join(str(v) for v in row[:8] if v is not None)
        score = metric_score(row_label)
        if score >= 12:
            observations = []
            for dr, dc, date_value in date_cells:
                if dc < len(row):
                    value = number_or_none(str(row[dc] or ""))
                    if value is not None and dr != r:
                        observations.append((date_value, value, r, dc, dr, dc))
            if observations:
                candidates.append((score, "dates_across_columns", row_label, observations))

    # Column-oriented metric: metric label above a column, dates down rows.
    for c in range(max_columns):
        column_label = " ".join(
            str(rows[r][c]) for r in range(min(len(rows), 30))
            if c < len(rows[r]) and rows[r][c] is not None
        )
        score = metric_score(column_label)
        if score >= 12:
            observations = []
            for dr, dc, date_value in date_cells:
                if c < len(rows[dr]):
                    value = number_or_none(str(rows[dr][c] or ""))
                    if value is not None and dc != c:
                        observations.append((date_value, value, dr, c, dr, dc))
            if observations:
                candidates.append((score, "dates_down_rows", column_label, observations))

    # Explicit fallback for the common BoE layout: find a cell containing
    # "1-year ahead expectations" and pair it with a nearby "Single month"
    # row/column before reading the intersecting date series.
    if not candidates:
        for r, row in enumerate(rows):
            for c, value in enumerate(row):
                label = normalize_excel_label(value)
                if not (re.search(r"(?:1|one) year ahead", label) and "expect" in label):
                    continue
                for rr in range(max(0, r - 4), min(len(rows), r + 6)):
                    nearby = normalize_excel_label(" ".join(
                        str(v) for v in rows[rr] if v is not None
                    ))
                    if "single month" not in nearby or "3 month" in nearby:
                        continue
                    observations = []
                    for dr, dc, date_value in date_cells:
                        if dc < len(rows[rr]):
                            numeric = number_or_none(str(rows[rr][dc] or ""))
                            if numeric is not None:
                                observations.append((date_value, numeric, rr, dc, dr, dc))
                    if observations:
                        candidates.append((20, "dates_across_columns", nearby, observations))

    if not candidates:
        preview = [
            [clean_cell(str(v or "")) for v in row[:12]]
            for row in rows[:30]
        ]
        raise RuntimeError(
            "Could not identify the DMP '1-year ahead expectations - Single month' "
            f"series. First rows: {preview}"
        )

    candidates.sort(key=lambda item: (item[0], len(item[3])), reverse=True)
    score, orientation, selected_label, observations = candidates[0]
    by_month = {}
    extracted_rows = []
    for date_value, value, vr, vc, dr, dc in observations:
        by_month[date_value] = {
            "date": date_value,
            "value": value,
            "source_url": source_url,
            "source_sheet": sheet.title,
            "source_cell": f"R{vr + 1}C{vc + 1}",
            "measure": "DMP 1-year ahead CPI inflation expectations - Single month",
        }
        extracted_rows.append({
            "date": date_value,
            "value": value,
            "date_cell": f"R{dr + 1}C{dc + 1}",
            "value_cell": f"R{vr + 1}C{vc + 1}",
        })
    points = sorted(by_month.values(), key=lambda point: point["date"])
    debug_payload = {
        "source_url": source_url,
        "sheet": sheet.title,
        "orientation": orientation,
        "selected_label": selected_label,
        "score": score,
        "date_cell_count": len(date_cells),
        "observation_count": len(points),
        "latest_observations": sorted(extracted_rows, key=lambda x: x["date"])[-24:],
    }
    return points, debug_payload


def update_dmp_inflation(database: dict[str, Any]) -> tuple[int, int]:
    reference_month, page_url, xlsx_url, response = download_latest_dmp_excel()
    print(
        f"[BoE DMP] reference_month={reference_month} xlsx={xlsx_url}",
        flush=True,
    )
    workbook = load_workbook(
        BytesIO(response.content), read_only=True, data_only=True
    )
    points, debug_payload = extract_dmp_single_month_cpi(workbook, xlsx_url)
    debug_payload.update({
        "release_page_url": page_url,
        "reference_month": reference_month,
        "content_type": response.headers.get("content-type"),
        "content_bytes": len(response.content),
        "workbook_sheets": workbook.sheetnames,
    })
    save_debug("boe_dmp_single_month_excel_debug.json", debug_payload)
    debug_print("boe_dmp_single_month_excel", debug_payload)
    print(
        f"[BoE DMP] extracted={len(points)} latest={points[-1]}",
        flush=True,
    )
    return merge(
        database,
        "ukbfftin",
        points,
        replace_source_range=True,
    )

def main() -> None:
    started_at = time.monotonic()
    print("[START] Update UK macro data", flush=True)
    database = json.loads(DATA_FILE.read_text(encoding="utf-8"))
    logs = []

    for series_id, (dataset, cdid, path, month_shift) in ONS.items():
        print(f"[ONS] updating {series_id} ({cdid})", flush=True)
        try:
            points = ons_series(dataset, cdid, path)
            if month_shift:
                points = [
                    {**point, "date": shift_month(point["date"], month_shift)}
                    for point in points
                ]
            logs.append((series_id, *merge(
                database, series_id, points, replace_source_range=True
            )))
        except Exception as error:
            logs.append((series_id, "ERROR", str(error)))

    for series_id, (dataset, cdid, path) in LEVELS.items():
        print(f"[ONS LEVEL] updating {series_id} ({cdid})", flush=True)
        try:
            points = year_over_year(ons_series(dataset, cdid, path))
            logs.append((series_id, *merge(
                database, series_id, points, replace_source_range=True
            )))
        except Exception as error:
            logs.append((series_id, "ERROR", str(error)))

    print("[S&P PMI] update started", flush=True)
    try:
        pmi_results = update_sp_global_pmi(database)
        for pmi_id, result in pmi_results.items():
            logs.append((pmi_id, *result))
    except Exception as error:
        logs.append(("sp_global_pmi", "ERROR", str(error)))

    updates = [
        ("ukrvayoy", lambda: update_retail(database)),
        ("ukcci", lambda: update_gfk(database)),
        ("ukbfftin", lambda: update_dmp_inflation(database)),
    ]

    for name, update_function in updates:
        print(f"[EXTRA] updating {name}", flush=True)
        try:
            logs.append((name, *update_function()))
        except Exception as error:
            logs.append((name, "ERROR", str(error)))

    database["generated_at"] = datetime.now(timezone.utc).isoformat()
    DATA_FILE.write_text(
        json.dumps(database, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    print("\n[UPDATE SUMMARY]", flush=True)
    for entry in logs:
        print(*entry, flush=True)
    elapsed = time.monotonic() - started_at
    print(f"\n[DONE] elapsed_seconds={elapsed:.1f}", flush=True)
    print(f"Debug files saved under: {DEBUG_DIR}", flush=True)


if __name__ == "__main__":
    main()
