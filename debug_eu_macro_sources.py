#!/usr/bin/env python3
"""Debug official sources for the indicators listed in EU_ECON.xlsx.

This script is intentionally read-only: it downloads recent observations, compares
up to four populated spreadsheet cells, and writes machine-readable diagnostics to
``debug/eu_macro_sources``. It does not update production JSON.

Designed for GitHub Actions. Required packages: requests, openpyxl.
Optional secrets are not required for Eurostat, INE, INSEE, or Bundesbank.
``GENESIS_TOKEN`` can be supplied later for Destatis GENESIS requests.
"""
from __future__ import annotations

import argparse
import csv
import io
import json
import math
import os
import re
import sys
import time
import xml.etree.ElementTree as ET
from dataclasses import asdict, dataclass
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any, Callable

import requests
from openpyxl import load_workbook

VERSION = "2026-07-29-v1"
DEFAULT_XLSX = Path("EU_ECON.xlsx")
DEFAULT_OUT = Path("debug/eu_macro_sources")
USER_AGENT = "Mozilla/5.0 (compatible; EUMacroSourceDebugger/1.0; GitHub-Actions)"
SESSION = requests.Session()
SESSION.headers.update({"User-Agent": USER_AGENT, "Accept-Language": "en,en-US;q=0.9"})

EUROSTAT = "https://ec.europa.eu/eurostat/api/dissemination/statistics/1.0/data"
INE = "https://servicios.ine.es/wstempus/js/ES/DATOS_SERIE"
INSEE = "https://api.insee.fr/series/BDM/V1/data/SERIES_BDM"
BUNDESBANK = "https://api.statistiken.bundesbank.de/rest/data"


@dataclass
class Point:
    period: str
    value: float
    source_url: str
    status: str = ""


@dataclass
class Candidate:
    name: str
    provider: str
    official: bool
    fetcher: str
    args: dict[str, Any]
    definition_note: str = ""


def log(message: str) -> None:
    print(message, flush=True)


def get(url: str, *, params: dict[str, Any] | None = None, headers: dict[str, str] | None = None) -> requests.Response:
    last: Exception | None = None
    for attempt in range(3):
        try:
            response = SESSION.get(url, params=params, headers=headers, timeout=45)
            log(f"[HTTP] {response.status_code} {response.url} bytes={len(response.content)}")
            if response.status_code < 400:
                return response
            if response.status_code not in {429, 500, 502, 503, 504}:
                response.raise_for_status()
        except requests.RequestException as exc:
            last = exc
        if attempt < 2:
            time.sleep(2 ** attempt * 2)
    if last:
        raise last
    response.raise_for_status()
    return response


def number(value: Any) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value) if math.isfinite(float(value)) else None
    text = str(value).strip().replace("%", "").replace(" ", "").replace(",", ".")
    if not text or text in {"-", "..", ":", "NA", "N/A"}:
        return None
    match = re.fullmatch(r"[+-]?\d+(?:\.\d+)?", text)
    return float(text) if match else None


def month_key(value: Any) -> str | None:
    if isinstance(value, (datetime, date)):
        return f"{value.year:04d}-{value.month:02d}"
    text = str(value or "").strip()
    match = re.match(r"^(20\d{2})[-/]?(0[1-9]|1[0-2])", text)
    if match:
        return f"{match.group(1)}-{match.group(2)}"
    q = re.match(r"^(20\d{2})[- ]?Q([1-4])$", text, re.I)
    if q:
        return f"{q.group(1)}-Q{q.group(2)}"
    return None


EMBEDDED_TARGETS: list[dict[str, Any]] = [
    {
        "row": 2,
        "section": "通膨",
        "label": "西Core CPI",
        "description": "Spain CPI Core YoY",
        "declared_source": "INE",
        "possible_code": "IPC208611",
        "expected": [
            {
                "period": "2026-06",
                "value": 2.9,
                "cell": "R2C7"
            },
            {
                "period": "2026-05",
                "value": 3.0,
                "cell": "R2C8"
            },
            {
                "period": "2026-04",
                "value": 2.8,
                "cell": "R2C9"
            }
        ]
    },
    {
        "row": 3,
        "section": "通膨",
        "label": "法 Core CPI",
        "description": "France CPI All Ex Energy YoY",
        "declared_source": "INSEE National Statistics Offi",
        "possible_code": "001768579",
        "expected": [
            {
                "period": "2026-06",
                "value": 1.00498,
                "cell": "R3C7"
            },
            {
                "period": "2026-05",
                "value": 1.25824,
                "cell": "R3C8"
            },
            {
                "period": "2026-04",
                "value": 1.15814,
                "cell": "R3C9"
            }
        ]
    },
    {
        "row": 4,
        "section": "通膨",
        "label": "德 Core CPI",
        "description": "Germany CPI Overall Index excl",
        "declared_source": "German Federal Statistical Off",
        "possible_code": "DE CPI ex Food&Energy Y",
        "expected": [
            {
                "period": "2026-06",
                "value": 2.45139,
                "cell": "R4C7"
            },
            {
                "period": "2026-05",
                "value": 2.54022,
                "cell": "R4C8"
            },
            {
                "period": "2026-04",
                "value": 2.29007,
                "cell": "R4C9"
            }
        ]
    },
    {
        "row": 5,
        "section": "通膨",
        "label": "歐 Core CPI",
        "description": "Eurostat Eurozone Core MUICP Y",
        "declared_source": "Eurostat",
        "possible_code": "YoY",
        "expected": [
            {
                "period": "2026-06",
                "value": 2.4,
                "cell": "R5C7"
            },
            {
                "period": "2026-05",
                "value": 2.6,
                "cell": "R5C8"
            },
            {
                "period": "2026-04",
                "value": 2.2,
                "cell": "R5C9"
            }
        ]
    },
    {
        "row": 6,
        "section": "失業率",
        "label": "法 失業率",
        "description": "France Unemployment Rate ILO M",
        "declared_source": "INSEE National Statistics Offi",
        "possible_code": "001688527",
        "expected": []
    },
    {
        "row": 7,
        "section": "失業率",
        "label": "西 失業率",
        "description": "Spain Unemployment Rate",
        "declared_source": "INE",
        "possible_code": "EPA815",
        "expected": []
    },
    {
        "row": 8,
        "section": "失業率",
        "label": "德 Unemployment Rate SWDA",
        "description": "Germany Unemployment Rate SWDA",
        "declared_source": "Deutsche Bundesbank",
        "possible_code": "DE Unemployment Rate SWDA",
        "expected": [
            {
                "period": "2026-06",
                "value": 6.3,
                "cell": "R8C7"
            },
            {
                "period": "2026-05",
                "value": 6.3,
                "cell": "R8C8"
            },
            {
                "period": "2026-04",
                "value": 6.4,
                "cell": "R8C9"
            }
        ]
    },
    {
        "row": 9,
        "section": "失業率",
        "label": "歐 失業率",
        "description": "Eurostat Unemployment Eurozone",
        "declared_source": "Eurostat",
        "possible_code": "Eurozone",
        "expected": [
            {
                "period": "2026-05",
                "value": 6.2,
                "cell": "R9C8"
            },
            {
                "period": "2026-04",
                "value": 6.2,
                "cell": "R9C9"
            }
        ]
    },
    {
        "row": 10,
        "section": "其他就業",
        "label": "西 就業",
        "description": "Spain Registered Employed Tota",
        "declared_source": "Spanish Labour Ministry",
        "possible_code": "Employed Tot Net change MoM SA",
        "expected": [
            {
                "period": "2026-06",
                "value": 92.5299999999988,
                "cell": "R10C7"
            },
            {
                "period": "2026-05",
                "value": 63.7400000000016,
                "cell": "R10C8"
            },
            {
                "period": "2026-04",
                "value": 41.75,
                "cell": "R10C9"
            }
        ]
    },
    {
        "row": 11,
        "section": "其他就業",
        "label": "德 失業人口",
        "description": "Germany Unemployment Change SW",
        "declared_source": "Deutsche Bundesbank",
        "possible_code": "Unemploy. change",
        "expected": [
            {
                "period": "2026-06",
                "value": -1.0,
                "cell": "R11C7"
            },
            {
                "period": "2026-05",
                "value": -12.0,
                "cell": "R11C8"
            },
            {
                "period": "2026-04",
                "value": 18.0,
                "cell": "R11C9"
            }
        ]
    },
    {
        "row": 12,
        "section": "零售",
        "label": "西 零售",
        "description": "Spain Retail Sales Constant Pr",
        "declared_source": "INE",
        "possible_code": "ICM2522",
        "expected": [
            {
                "period": "2026-05",
                "value": -0.4,
                "cell": "R12C8"
            },
            {
                "period": "2026-04",
                "value": 0.2,
                "cell": "R12C9"
            }
        ]
    },
    {
        "row": 13,
        "section": "零售",
        "label": "德 零售",
        "description": "Germany Retail Sales Constant",
        "declared_source": "German Federal Statistical Off",
        "possible_code": "DE Rtl Sls Real SWDA MoM",
        "expected": [
            {
                "period": "2026-05",
                "value": 1.0,
                "cell": "R13C8"
            },
            {
                "period": "2026-04",
                "value": -0.2,
                "cell": "R13C9"
            }
        ]
    },
    {
        "row": 14,
        "section": "零售",
        "label": "歐 Real零售",
        "description": "Eurostat Retail Sales Eurozone",
        "declared_source": "Eurostat",
        "possible_code": "WDA YoY %",
        "expected": [
            {
                "period": "2026-05",
                "value": 1.6,
                "cell": "R14C8"
            },
            {
                "period": "2026-04",
                "value": 0.9,
                "cell": "R14C9"
            }
        ]
    },
    {
        "row": 15,
        "section": "工業",
        "label": "德 工業",
        "description": "Germany Industrial Production",
        "declared_source": "Bundesministerium fur Wirtscha",
        "possible_code": "Ind Prod YoY NSA WDA",
        "expected": [
            {
                "period": "2026-05",
                "value": 0.0,
                "cell": "R15C8"
            },
            {
                "period": "2026-04",
                "value": -0.8762322015334,
                "cell": "R15C9"
            }
        ]
    },
    {
        "row": 16,
        "section": "消費者信心",
        "label": "法 信心",
        "description": "France Consumer Confidence Ove",
        "declared_source": "INSEE National Statistics Offi",
        "possible_code": "001587668",
        "expected": [
            {
                "period": "2026-06",
                "value": 84.0,
                "cell": "R16C7"
            },
            {
                "period": "2026-05",
                "value": 82.0,
                "cell": "R16C8"
            },
            {
                "period": "2026-04",
                "value": 84.0,
                "cell": "R16C9"
            }
        ]
    },
    {
        "row": 17,
        "section": "消費者信心",
        "label": "德 GfK Consumer Confidence",
        "description": "GfK Consumer Confidence",
        "declared_source": "GfK SE",
        "possible_code": "GfK Confidence (No History) Co",
        "expected": [
            {
                "period": "2026-07",
                "value": -29.3,
                "cell": "R17C6"
            },
            {
                "period": "2026-06",
                "value": -29.7,
                "cell": "R17C7"
            },
            {
                "period": "2026-05",
                "value": -33.1,
                "cell": "R17C8"
            },
            {
                "period": "2026-04",
                "value": -28.1,
                "cell": "R17C9"
            }
        ]
    },
    {
        "row": 18,
        "section": "消費者信心",
        "label": "德信心 Current",
        "description": "ZEW Germany Assessment of Curr",
        "declared_source": "ZEW Zentrum fuer Europaeische",
        "possible_code": "Current Situation",
        "expected": [
            {
                "period": "2026-07",
                "value": -77.6,
                "cell": "R18C6"
            },
            {
                "period": "2026-06",
                "value": -81.0,
                "cell": "R18C7"
            },
            {
                "period": "2026-05",
                "value": -77.8,
                "cell": "R18C8"
            },
            {
                "period": "2026-04",
                "value": -73.7,
                "cell": "R18C9"
            }
        ]
    },
    {
        "row": 19,
        "section": "消費者信心",
        "label": "德信心 expect",
        "description": "ZEW Germany Expectation of Eco",
        "declared_source": "ZEW Zentrum fuer Europaeische",
        "possible_code": "Expectations",
        "expected": [
            {
                "period": "2026-07",
                "value": 26.3,
                "cell": "R19C6"
            },
            {
                "period": "2026-06",
                "value": 10.5,
                "cell": "R19C7"
            },
            {
                "period": "2026-05",
                "value": -10.2,
                "cell": "R19C8"
            },
            {
                "period": "2026-04",
                "value": -17.2,
                "cell": "R19C9"
            }
        ]
    },
    {
        "row": 20,
        "section": "製造業",
        "label": "德 製造業PMI",
        "description": "S&P Global/BME Germany Manufac",
        "declared_source": "S&P Global",
        "possible_code": "",
        "expected": [
            {
                "period": "2026-07",
                "value": 52.2,
                "cell": "R20C6"
            },
            {
                "period": "2026-06",
                "value": 50.3,
                "cell": "R20C7"
            },
            {
                "period": "2026-05",
                "value": 50.1,
                "cell": "R20C8"
            },
            {
                "period": "2026-04",
                "value": 51.4,
                "cell": "R20C9"
            }
        ]
    },
    {
        "row": 21,
        "section": "製造業",
        "label": "法 製造業PMI",
        "description": "France Manufacturing PMI SA",
        "declared_source": "S&P Global",
        "possible_code": "",
        "expected": [
            {
                "period": "2026-07",
                "value": 50.0,
                "cell": "R21C6"
            },
            {
                "period": "2026-06",
                "value": 51.2,
                "cell": "R21C7"
            },
            {
                "period": "2026-05",
                "value": 49.7,
                "cell": "R21C8"
            },
            {
                "period": "2026-04",
                "value": 52.8,
                "cell": "R21C9"
            }
        ]
    },
    {
        "row": 22,
        "section": "製造業",
        "label": "法 製造業信心",
        "description": "France Business Confidence Man",
        "declared_source": "INSEE National Statistics Offi",
        "possible_code": "001585934",
        "expected": [
            {
                "period": "2026-07",
                "value": 101.3,
                "cell": "R22C6"
            },
            {
                "period": "2026-06",
                "value": 100.2,
                "cell": "R22C7"
            },
            {
                "period": "2026-05",
                "value": 102.2,
                "cell": "R22C8"
            },
            {
                "period": "2026-04",
                "value": 100.2,
                "cell": "R22C9"
            }
        ]
    },
    {
        "row": 23,
        "section": "製造業",
        "label": "西 製造業PMI",
        "description": "Spain Manufacturing PMI SA",
        "declared_source": "S&P Global",
        "possible_code": "",
        "expected": [
            {
                "period": "2026-06",
                "value": 49.7,
                "cell": "R23C7"
            },
            {
                "period": "2026-05",
                "value": 51.2,
                "cell": "R23C8"
            },
            {
                "period": "2026-04",
                "value": 51.7,
                "cell": "R23C9"
            }
        ]
    },
    {
        "row": 24,
        "section": "服務業",
        "label": "德 服務業PMI",
        "description": "Germany Services PMI Business",
        "declared_source": "S&P Global",
        "possible_code": "",
        "expected": [
            {
                "period": "2026-07",
                "value": 49.6,
                "cell": "R24C6"
            },
            {
                "period": "2026-06",
                "value": 48.6,
                "cell": "R24C7"
            },
            {
                "period": "2026-05",
                "value": 48.1,
                "cell": "R24C8"
            },
            {
                "period": "2026-04",
                "value": 46.9,
                "cell": "R24C9"
            }
        ]
    },
    {
        "row": 25,
        "section": "服務業",
        "label": "法 服務業PMI",
        "description": "France Services PMI SA",
        "declared_source": "S&P Global",
        "possible_code": "",
        "expected": [
            {
                "period": "2026-07",
                "value": 49.8,
                "cell": "R25C6"
            },
            {
                "period": "2026-06",
                "value": 46.8,
                "cell": "R25C7"
            },
            {
                "period": "2026-05",
                "value": 44.3,
                "cell": "R25C8"
            },
            {
                "period": "2026-04",
                "value": 46.5,
                "cell": "R25C9"
            }
        ]
    },
    {
        "row": 26,
        "section": "服務業",
        "label": "西 服務業PMI",
        "description": "Spain Services PMI Business Ac",
        "declared_source": "S&P Global",
        "possible_code": "",
        "expected": [
            {
                "period": "2026-06",
                "value": 54.2,
                "cell": "R26C7"
            },
            {
                "period": "2026-05",
                "value": 50.1,
                "cell": "R26C8"
            },
            {
                "period": "2026-04",
                "value": 47.9,
                "cell": "R26C9"
            }
        ]
    },
    {
        "row": 27,
        "section": "企業信心",
        "label": "法 企業信心",
        "description": "France Business Confidence Com",
        "declared_source": "INSEE National Statistics Offi",
        "possible_code": "001565530",
        "expected": [
            {
                "period": "2026-07",
                "value": 97.2,
                "cell": "R27C6"
            },
            {
                "period": "2026-06",
                "value": 95.0,
                "cell": "R27C7"
            },
            {
                "period": "2026-05",
                "value": 93.9,
                "cell": "R27C8"
            },
            {
                "period": "2026-04",
                "value": 94.1,
                "cell": "R27C9"
            }
        ]
    },
    {
        "row": 28,
        "section": "企業信心",
        "label": "德 企業信心",
        "description": "Ifo Pan Germany Business Clima",
        "declared_source": "IFO Institute - Institut fuer",
        "possible_code": "Business Climate",
        "expected": [
            {
                "period": "2026-07",
                "value": 86.59582,
                "cell": "R28C6"
            },
            {
                "period": "2026-06",
                "value": 85.7,
                "cell": "R28C7"
            },
            {
                "period": "2026-05",
                "value": 85.0,
                "cell": "R28C8"
            },
            {
                "period": "2026-04",
                "value": 84.5,
                "cell": "R28C9"
            }
        ]
    },
    {
        "row": 31,
        "section": "GDP",
        "label": "德 GDP",
        "description": "Germany GDP Chain Linked Pan G",
        "declared_source": "German Federal Statistical Off",
        "possible_code": "YoY",
        "expected": [
            {
                "period": "2026-07",
                "value": 0.5,
                "cell": "R31C6"
            },
            {
                "period": "2026-06",
                "value": 0.5,
                "cell": "R31C7"
            },
            {
                "period": "2026-05",
                "value": 0.3,
                "cell": "R31C8"
            },
            {
                "period": "2026-04",
                "value": 0.0,
                "cell": "R31C9"
            }
        ]
    },
    {
        "row": 32,
        "section": "GDP",
        "label": "西 GDP",
        "description": "Spain GDP SA Chained Linked at",
        "declared_source": "INE",
        "possible_code": "CNTR4892",
        "expected": [
            {
                "period": "2026-07",
                "value": 2.7126,
                "cell": "R32C6"
            },
            {
                "period": "2026-06",
                "value": 2.6461,
                "cell": "R32C7"
            },
            {
                "period": "2026-05",
                "value": 2.703,
                "cell": "R32C8"
            },
            {
                "period": "2026-04",
                "value": 2.8792,
                "cell": "R32C9"
            }
        ]
    },
    {
        "row": 33,
        "section": "GDP",
        "label": "法GDP",
        "description": "France GDP Chain Linked Prices",
        "declared_source": "INSEE National Statistics Offi",
        "possible_code": "GDP YoY",
        "expected": [
            {
                "period": "2026-07",
                "value": 0.87354,
                "cell": "R33C6"
            },
            {
                "period": "2026-06",
                "value": 1.1,
                "cell": "R33C7"
            },
            {
                "period": "2026-05",
                "value": 0.8,
                "cell": "R33C8"
            },
            {
                "period": "2026-04",
                "value": 0.8,
                "cell": "R33C9"
            }
        ]
    },
    {
        "row": 34,
        "section": "GDP",
        "label": "歐GDP",
        "description": "Euro Area Gross Domestic Produ",
        "declared_source": "Eurostat",
        "possible_code": "EA GDP YoY",
        "expected": [
            {
                "period": "2026-07",
                "value": 0.5,
                "cell": "R34C6"
            },
            {
                "period": "2026-06",
                "value": 1.1,
                "cell": "R34C7"
            },
            {
                "period": "2026-05",
                "value": 1.2,
                "cell": "R34C8"
            },
            {
                "period": "2026-04",
                "value": 1.4,
                "cell": "R34C9"
            }
        ]
    }
]


def read_targets(path: Path) -> list[dict[str, Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise RuntimeError("EU_ECON workbook is empty")
    header = rows[0]
    date_columns = [(index, month_key(value)) for index, value in enumerate(header) if index >= 5 and month_key(value)]
    output: list[dict[str, Any]] = []
    section = ""
    for row_no, row in enumerate(rows[1:], 2):
        if row and row[0]:
            section = str(row[0]).strip()
        label = str(row[1] or "").strip() if len(row) > 1 else ""
        if not label:
            continue
        expected = []
        for col, period in date_columns:
            value = number(row[col] if col < len(row) else None)
            if value is not None:
                expected.append({"period": period, "value": value, "cell": f"R{row_no}C{col+1}"})
        output.append({
            "row": row_no,
            "section": section,
            "label": label,
            "description": str(row[2] or "").strip(),
            "declared_source": str(row[3] or "").strip(),
            "possible_code": str(row[4] or "").strip(),
            "expected": expected[:4],
        })
    return output


def flatten_index(position: int, sizes: list[int]) -> list[int]:
    coords = [0] * len(sizes)
    for i in range(len(sizes) - 1, -1, -1):
        coords[i] = position % sizes[i]
        position //= sizes[i]
    return coords


def eurostat(dataset: str, filters: dict[str, str]) -> list[Point]:
    params = {"lang": "EN", "format": "JSON", **filters}
    response = get(f"{EUROSTAT}/{dataset}", params=params)
    payload = response.json()
    ids = payload.get("id", [])
    sizes = payload.get("size", [])
    values = payload.get("value", {})
    if not ids or not sizes or "time" not in ids:
        raise RuntimeError(f"Unexpected Eurostat JSON-stat structure for {dataset}")
    categories: dict[str, list[str]] = {}
    for dim in ids:
        index = payload["dimension"][dim]["category"]["index"]
        if isinstance(index, dict):
            ordered = [None] * len(index)
            for code, pos in index.items():
                ordered[int(pos)] = code
            categories[dim] = ordered
        else:
            categories[dim] = list(index)
    points = []
    for raw_pos, raw_value in values.items():
        coords = flatten_index(int(raw_pos), sizes)
        record = {dim: categories[dim][coords[i]] for i, dim in enumerate(ids)}
        period = month_key(record.get("time"))
        value = number(raw_value)
        if period and value is not None:
            points.append(Point(period, value, response.url, str(payload.get("status", {}).get(raw_pos, ""))))
    return dedupe(points)


def ine_series(code: str) -> list[Point]:
    response = get(f"{INE}/{code}", params={"nult": 18})
    payload = response.json()
    rows = payload[0].get("Data", []) if isinstance(payload, list) and payload else payload.get("Data", [])
    points = []
    for row in rows:
        period = month_key(row.get("Fecha"))
        if not period and row.get("Anyo") and row.get("FK_Periodo"):
            match = re.search(r"(\d{1,2})$", str(row.get("FK_Periodo")))
            if match:
                period = f"{int(row['Anyo']):04d}-{int(match.group(1)):02d}"
        value = number(row.get("Valor"))
        if period and value is not None:
            points.append(Point(period, value, response.url, str(row.get("T3_TipoDato", ""))))
    if not points:
        raise RuntimeError(f"INE {code} returned no observations")
    return dedupe(points)


def insee_series(idbank: str) -> list[Point]:
    response = get(f"{INSEE}/{idbank}", params={"lastNObservations": 18})
    root = ET.fromstring(response.content)
    points = []
    for obs in root.iter():
        if not obs.tag.endswith("Obs"):
            continue
        period = month_key(obs.attrib.get("TIME_PERIOD") or obs.attrib.get("timePeriod"))
        value = number(obs.attrib.get("OBS_VALUE") or obs.attrib.get("obsValue"))
        if period and value is not None:
            points.append(Point(period, value, response.url, obs.attrib.get("OBS_STATUS", "")))
    if not points:
        raise RuntimeError(f"INSEE idbank {idbank} returned no observations")
    return dedupe(points)


def bundesbank(flow: str, key: str) -> list[Point]:
    response = get(
        f"{BUNDESBANK}/{flow}/{key}",
        params={"format": "sdmx_csv", "lang": "en", "startPeriod": "2025-01"},
        headers={"Accept": "text/csv"},
    )
    text = response.content.decode("utf-8-sig", "replace")
    reader = csv.DictReader(io.StringIO(text))
    points = []
    for row in reader:
        period = month_key(row.get("TIME_PERIOD"))
        value = number(row.get("OBS_VALUE"))
        if period and value is not None:
            points.append(Point(period, value, response.url, row.get("OBS_STATUS", "")))
    if not points:
        raise RuntimeError(f"Bundesbank {flow}/{key} returned no observations")
    return dedupe(points)


def dedupe(points: list[Point]) -> list[Point]:
    by_period = {point.period: point for point in points}
    return [by_period[key] for key in sorted(by_period)]


# Candidate definitions. Multiple official candidates are deliberately retained for
# ambiguous vendor labels; the report ranks them by overlap and error instead of
# silently selecting a similar but definitionally different series.
C: dict[str, list[Candidate]] = {
    "西Core CPI": [Candidate("INE IPC208611", "INE", True, "ine", {"code": "IPC208611"}, "National CPI core, year-on-year")],
    "法 Core CPI": [Candidate("INSEE 001768579", "INSEE", True, "insee", {"idbank": "001768579"}, "French national CPI excluding energy, year-on-year")],
    "歐 Core CPI": [Candidate("Euro area HICP excluding energy and food", "Eurostat", True, "eurostat", {"dataset": "prc_hicp_manr", "filters": {"geo": "EA20", "unit": "RCH_A", "coicop": "TOT_X_NRG_FOOD"}}, "HICP, not national CPI")],
    "法 失業率": [Candidate("INSEE 001688527", "INSEE", True, "insee", {"idbank": "001688527"}, "ILO unemployment rate; verify monthly-vs-quarterly vintage")],
    "西 失業率": [Candidate("INE EPA815", "INE", True, "ine", {"code": "EPA815"}, "Spanish Labour Force Survey unemployment rate")],
    "歐 失業率": [Candidate("Euro area unemployment rate", "Eurostat", True, "eurostat", {"dataset": "une_rt_m", "filters": {"geo": "EA20", "age": "Y15-74", "sex": "T", "s_adj": "SA", "unit": "PC_ACT"}}, "Seasonally adjusted, ages 15-74")],
    "西 零售": [Candidate("INE ICM2522", "INE", True, "ine", {"code": "ICM2522"}, "Retail trade constant-price series")],
    "法 信心": [Candidate("INSEE 001587668", "INSEE", True, "insee", {"idbank": "001587668"}, "Household confidence synthetic index")],
    "法 製造業信心": [Candidate("INSEE 001585934", "INSEE", True, "insee", {"idbank": "001585934"}, "Manufacturing business climate")],
    "法 企業信心": [Candidate("INSEE 001565530", "INSEE", True, "insee", {"idbank": "001565530"}, "All-sector business climate")],
    "德 GDP": [Candidate("Germany real GDP YoY", "Eurostat", True, "eurostat", {"dataset": "namq_10_gdp", "filters": {"geo": "DE", "na_item": "B1GQ", "unit": "CLV_PCH_SM", "s_adj": "SCA"}}, "Quarterly real GDP change from same quarter previous year")],
    "西 GDP": [Candidate("INE CNTR4892", "INE", True, "ine", {"code": "CNTR4892"}, "Quarterly chained-volume GDP year-on-year")],
    "法GDP": [Candidate("France real GDP YoY", "Eurostat", True, "eurostat", {"dataset": "namq_10_gdp", "filters": {"geo": "FR", "na_item": "B1GQ", "unit": "CLV_PCH_SM", "s_adj": "SCA"}}, "Quarterly real GDP change from same quarter previous year")],
    "歐GDP": [Candidate("Euro area real GDP YoY", "Eurostat", True, "eurostat", {"dataset": "namq_10_gdp", "filters": {"geo": "EA20", "na_item": "B1GQ", "unit": "CLV_PCH_SM", "s_adj": "SCA"}}, "Quarterly real GDP change from same quarter previous year")],
    "德 Unemployment Rate SWDA": [Candidate("Germany unemployment rate", "Eurostat", True, "eurostat", {"dataset": "une_rt_m", "filters": {"geo": "DE", "age": "Y15-74", "sex": "T", "s_adj": "SA", "unit": "PC_ACT"}}, "ILO concept; may differ from German registered-unemployment rate")],
    "德 零售": [Candidate("Germany retail volume MoM", "Eurostat", True, "eurostat", {"dataset": "sts_trtu_m", "filters": {"geo": "DE", "unit": "PCH_PRE", "s_adj": "SCA", "nace_r2": "G47"}}, "Retail volume, seasonally/calendar adjusted, month-on-month")],
    "歐 Real零售": [Candidate("Euro area retail volume YoY", "Eurostat", True, "eurostat", {"dataset": "sts_trtu_m", "filters": {"geo": "EA20", "unit": "PCH_SM", "s_adj": "CA", "nace_r2": "G47"}}, "Retail volume, calendar adjusted, year-on-year")],
    "德 工業": [Candidate("Germany industrial production YoY", "Eurostat", True, "eurostat", {"dataset": "sts_inpr_m", "filters": {"geo": "DE", "unit": "PCH_SM", "s_adj": "CA", "nace_r2": "B-D"}}, "Production index, calendar adjusted, year-on-year")],
}

MANUAL = {
    "西 就業": ("Spanish Ministry of Inclusion", "Official producer; exact vendor transformation is net monthly change, seasonally adjusted. Add a stable ministry CSV/XLSX URL before automation."),
    "德 失業人口": ("Bundesagentur fuer Arbeit / Bundesbank", "Official data exist, but the workbook has no Bundesbank time-series key. Supply the BBK flow/key to avoid picking the wrong unemployment concept."),
    "德 Core CPI": ("Destatis GENESIS", "National CPI excluding food and energy requires a precise GENESIS table/selection. GENESIS supports JSON and XLSX/CSV downloads; set GENESIS_TOKEN after confirming the table code."),
    "德 GfK Consumer Confidence": ("GfK", "Producer release, not a government/open statistical API; history is normally in press releases or licensed feeds."),
    "德信心 Current": ("ZEW", "Producer release; exact monthly history generally comes from ZEW releases, not an open JSON API."),
    "德信心 expect": ("ZEW", "Producer release; exact monthly history generally comes from ZEW releases, not an open JSON API."),
    "德 製造業PMI": ("S&P Global", "Official producer is S&P Global, but PMI history is proprietary; public press releases can validate recent releases only."),
    "法 製造業PMI": ("S&P Global", "Official producer is S&P Global, but PMI history is proprietary; public press releases can validate recent releases only."),
    "西 製造業PMI": ("S&P Global", "Official producer is S&P Global, but PMI history is proprietary; public press releases can validate recent releases only."),
    "德 服務業PMI": ("S&P Global", "Official producer is S&P Global, but PMI history is proprietary; public press releases can validate recent releases only."),
    "法 服務業PMI": ("S&P Global", "Official producer is S&P Global, but PMI history is proprietary; public press releases can validate recent releases only."),
    "西 服務業PMI": ("S&P Global", "Official producer is S&P Global, but PMI history is proprietary; public press releases can validate recent releases only."),
    "德 企業信心": ("ifo Institute", "Producer release; public releases validate recent values, but the exact seasonally adjusted history/download endpoint must be confirmed."),
}

FETCHERS: dict[str, Callable[..., list[Point]]] = {
    "eurostat": eurostat,
    "ine": ine_series,
    "insee": insee_series,
    "bundesbank": bundesbank,
}


def compare(expected: list[dict[str, Any]], actual: list[Point], tolerance: float) -> dict[str, Any]:
    actual_by_period = {p.period: p for p in actual}
    rows = []
    for item in expected:
        point = actual_by_period.get(item["period"])
        difference = None if point is None else point.value - item["value"]
        rows.append({
            **item,
            "actual": None if point is None else point.value,
            "difference": difference,
            "match": difference is not None and abs(difference) <= tolerance,
            "source_url": None if point is None else point.source_url,
        })
    matched_periods = sum(row["actual"] is not None for row in rows)
    exact = sum(row["match"] for row in rows)
    mae = None
    diffs = [abs(row["difference"]) for row in rows if row["difference"] is not None]
    if diffs:
        mae = sum(diffs) / len(diffs)
    return {"rows": rows, "overlap": matched_periods, "matches": exact, "mae": mae}


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--xlsx",
        type=Path,
        default=None,
        help="Optional EU_ECON workbook. If omitted or missing, use the reference values embedded in this script.",
    )
    parser.add_argument("--out", type=Path, default=DEFAULT_OUT)
    parser.add_argument("--tolerance", type=float, default=0.051, help="absolute comparison tolerance")
    args = parser.parse_args()
    args.out.mkdir(parents=True, exist_ok=True)
    if args.xlsx is not None and args.xlsx.exists():
        targets = read_targets(args.xlsx)
        reference_source = str(args.xlsx)
        log(f"[REFERENCE] Loaded workbook: {args.xlsx}")
    else:
        targets = EMBEDDED_TARGETS
        reference_source = "embedded EU_ECON reference values"
        if args.xlsx is not None:
            log(f"[REFERENCE] Workbook not found: {args.xlsx}; using embedded reference values")
        else:
            log("[REFERENCE] Using embedded EU_ECON reference values")
    report: dict[str, Any] = {
        "script_version": VERSION,
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "workbook": reference_source,
        "tolerance": args.tolerance,
        "results": [],
    }
    for target in targets:
        label = target["label"]
        log(f"\n=== {label} ===")
        item = {**target, "candidates": []}
        candidates = C.get(label, [])
        if not candidates:
            provider, note = MANUAL.get(label, (target["declared_source"] or "Unknown", "No exact official machine-readable mapping has been confirmed."))
            item.update({"status": "NEEDS_MAPPING", "provider": provider, "note": note})
        else:
            for candidate in candidates:
                result = {**asdict(candidate)}
                try:
                    points = FETCHERS[candidate.fetcher](**candidate.args)
                    result["latest_points"] = [asdict(p) for p in points[-12:]]
                    result["comparison"] = compare(target["expected"], points, args.tolerance)
                    result["status"] = "OK"
                except Exception as exc:
                    result["status"] = "ERROR"
                    result["error"] = f"{type(exc).__name__}: {exc}"
                    log(f"[ERROR] {candidate.name}: {result['error']}")
                item["candidates"].append(result)
            successful = [x for x in item["candidates"] if x["status"] == "OK"]
            successful.sort(key=lambda x: (-x["comparison"]["overlap"], x["comparison"]["mae"] if x["comparison"]["mae"] is not None else 1e99))
            if successful:
                best = successful[0]
                cmp = best["comparison"]
                item["selected_candidate"] = best["name"]
                item["status"] = "MATCH" if cmp["overlap"] and cmp["matches"] == cmp["overlap"] else ("MISMATCH" if cmp["overlap"] else "NO_PERIOD_OVERLAP")
            else:
                item["status"] = "FETCH_ERROR"
        report["results"].append(item)

    summary: dict[str, int] = {}
    for item in report["results"]:
        summary[item["status"]] = summary.get(item["status"], 0) + 1
    report["summary"] = summary
    json_path = args.out / "eu_macro_source_comparison.json"
    json_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")

    csv_path = args.out / "eu_macro_source_comparison.csv"
    with csv_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(["label", "status", "selected_candidate", "period", "EU_ECON", "official", "difference", "match", "source_url", "note"])
        for item in report["results"]:
            selected = next((x for x in item.get("candidates", []) if x.get("name") == item.get("selected_candidate")), None)
            if selected:
                for row in selected["comparison"]["rows"]:
                    writer.writerow([item["label"], item["status"], selected["name"], row["period"], row["value"], row["actual"], row["difference"], row["match"], row["source_url"], selected.get("definition_note", "")])
            else:
                writer.writerow([item["label"], item["status"], "", "", "", "", "", "", "", item.get("note", "")])

    log("\n[SUMMARY]")
    for status, count in sorted(summary.items()):
        log(f"{status}: {count}")
    log(f"JSON: {json_path}")
    log(f"CSV : {csv_path}")
    # Debugging should finish and upload reports even when individual sources fail.
    return 0


if __name__ == "__main__":
    sys.exit(main())
