#!/usr/bin/env python3
"""Validate structured EU macro sources against EU_ECON reference values.

This diagnostic intentionally prefers official machine-readable sources:
API/JSON/SDMX/CSV/XLSX/PX-Web. It does not use official news-release HTML
for values, and it excludes PMI because a free official structured feed is
not available.

Outputs:
  structured_source_comparison.json
  structured_source_comparison.csv
  structured_source_candidates.json
  structured_source_http_log.json

Run:
  python debug_eu_structured_sources.py --out debug/eu_structured_sources
"""
from __future__ import annotations

import argparse
import csv
import io
import json
import math
import re
import sys
import time
import traceback
import unicodedata
import xml.etree.ElementTree as ET
from dataclasses import asdict, dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Callable
from urllib.parse import unquote, urljoin

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

VERSION = "2026-07-30-structured-v3-value-pattern-and-raw-dump"
TIMEOUT = 45
USER_AGENT = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/126.0 Safari/537.36"

EXPECTED: dict[str, dict[str, float]] = {
    "西Core CPI": {"2026-06": 2.9, "2026-05": 3.0, "2026-04": 2.8},
    "法 Core CPI": {"2026-06": 1.00498, "2026-05": 1.25824, "2026-04": 1.15814},
    "德 Core CPI": {"2026-06": 2.45139, "2026-05": 2.54022, "2026-04": 2.29007},
    "歐 Core CPI": {"2026-06": 2.4, "2026-05": 2.6, "2026-04": 2.2},
    "西 失業率": {"2026-Q2": 9.87},
    "歐 失業率": {"2026-05": 6.2, "2026-04": 6.2},
    "西 就業": {"2026-06": 92.53, "2026-05": 63.74, "2026-04": 41.75},
    "西 零售": {"2026-05": -0.4, "2026-04": 0.2},
    "德 工業": {"2026-05": 0.0, "2026-04": -0.8762322015334},
    "德信心 Current": {"2026-07": -77.6, "2026-06": -81.0, "2026-05": -77.8, "2026-04": -73.7},
    "德信心 expect": {"2026-07": 26.3, "2026-06": 10.5, "2026-05": -10.2, "2026-04": -17.2},
    "德 企業信心": {"2026-07": 86.59582, "2026-06": 85.7, "2026-05": 85.0, "2026-04": 84.5},
    "德 GDP": {"2026-Q1": 0.5, "2025-Q4": 0.5, "2025-Q3": 0.3, "2025-Q2": 0.0},
}

HTTP_LOG: list[dict[str, Any]] = []
CANDIDATES: dict[str, Any] = {}


def log(message: str) -> None:
    print(message, flush=True)


def clean(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def norm(value: Any) -> str:
    text = unicodedata.normalize("NFKD", clean(value)).encode("ascii", "ignore").decode().lower()
    return re.sub(r"[^a-z0-9]+", " ", text).strip()


def number(value: Any) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value) if math.isfinite(float(value)) else None
    text = clean(value).replace("−", "-").replace("%", "")
    if not text or text.lower() in {"na", "nan", "null", "none", "..", ":", "-"}:
        return None
    text = re.sub(r"[^0-9,.-]", "", text)
    if text.count(",") == 1 and text.count(".") == 0:
        text = text.replace(",", ".")
    elif text.count(".") > 1 and text.count(",") == 0:
        text = text.replace(".", "")
    elif "," in text and "." in text:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    try:
        return float(text)
    except ValueError:
        return None


def period_key(value: Any) -> str | None:
    text = clean(value)
    if not text:
        return None
    q = re.search(r"(20\d{2})\D*[QqTt]([1-4])", text)
    if q:
        return f"{q.group(1)}-Q{q.group(2)}"
    q = re.search(r"[QqTt]([1-4])\D*(20\d{2})", text)
    if q:
        return f"{q.group(2)}-Q{q.group(1)}"
    m = re.search(r"(20\d{2})[-/. ](0?[1-9]|1[0-2])(?:\D|$)", text)
    if m:
        return f"{m.group(1)}-{int(m.group(2)):02d}"
    months = {name: i for i, name in enumerate(
        "january february march april may june july august september october november december".split(), 1)}
    months.update({name: i for i, name in enumerate(
        "enero febrero marzo abril mayo junio julio agosto septiembre octubre noviembre diciembre".split(), 1)})
    months.update({name: i for i, name in enumerate(
        "januar februar marz april mai juni juli august september oktober november dezember".split(), 1)})
    n = norm(text)
    year = re.search(r"\b(20\d{2})\b", n)
    if year:
        for name, month in months.items():
            if re.search(rf"\b{name}\b", n):
                return f"{year.group(1)}-{month:02d}"
    return None


@dataclass
class Point:
    period: str
    value: float
    source_url: str
    status: str = ""
    note: str = ""


@dataclass
class SourceTest:
    label: str
    source_name: str
    format: str
    official: bool
    definition: str
    fetcher: Callable[[], list[Point]]


def session() -> requests.Session:
    s = requests.Session()
    retry = Retry(total=2, connect=2, read=2, backoff_factor=1.0,
                  status_forcelist=(429, 500, 502, 503, 504), allowed_methods=("GET", "POST"))
    s.mount("https://", HTTPAdapter(max_retries=retry))
    s.headers.update({"User-Agent": USER_AGENT, "Accept": "*/*"})
    return s


HTTP = session()


def request(method: str, url: str, **kwargs: Any) -> requests.Response:
    started = time.time()
    try:
        response = HTTP.request(method, url, timeout=TIMEOUT, **kwargs)
        HTTP_LOG.append({"method": method, "url": response.url, "status": response.status_code,
                         "bytes": len(response.content), "seconds": round(time.time() - started, 3),
                         "content_type": response.headers.get("content-type", "")})
        response.raise_for_status()
        return response
    except Exception as exc:
        HTTP_LOG.append({"method": method, "url": url, "status": None,
                         "seconds": round(time.time() - started, 3), "error": f"{type(exc).__name__}: {exc}"})
        raise


def dedupe(points: list[Point]) -> list[Point]:
    out: dict[str, Point] = {}
    for point in points:
        if point.period:
            out[point.period] = point
    return [out[key] for key in sorted(out)]


# ---------- Generic structured-source parsers ----------

def ine_series(code: str) -> list[Point]:
    """Read an INE Tempus3 series using its actual observation date.

    FK_Periodo is a catalogue identifier, not a month/quarter number. The v1
    parser treated its trailing digits as a period and therefore shifted recent
    observations into old years. Fecha and NombrePeriodo are authoritative.
    """
    url = f"https://servicios.ine.es/wstempus/js/ES/DATOS_SERIE/{code}"
    response = request("GET", url, params={"nult": 40})
    payload = response.json()
    rows = payload[0].get("Data", []) if isinstance(payload, list) and payload else payload.get("Data", [])
    points: list[Point] = []
    for row in rows:
        period = period_key(row.get("Fecha"))
        if not period:
            year = int(row.get("Anyo", 0) or 0)
            name = clean(row.get("NombrePeriodo") or row.get("Periodo") or row.get("T3_Periodo"))
            q = re.search(r"(?:trimestre|quarter|t)\s*([1-4])|([1-4])\s*(?:trimestre|quarter|t)", name, re.I)
            if q and year:
                period = f"{year:04d}-Q{q.group(1) or q.group(2)}"
            else:
                period = period_key(f"{name} {year}")
        if not period:
            year = int(row.get("Anyo", 0) or 0)
            fk = clean(row.get("FK_Periodo") or row.get("T3_Periodo"))
            tail = re.search(r"(\d{1,2})$", fk)
            pcode = int(tail.group(1)) if tail else 0
            if year and 1 <= pcode <= 12:
                period = f"{year:04d}-{pcode:02d}"
            elif year and 19 <= pcode <= 22:
                period = f"{year:04d}-Q{pcode-18}"
        value = number(row.get("Valor"))
        if period and value is not None:
            points.append(Point(period, value, response.url, clean(row.get("T3_TipoDato")),
                                note=f"INE series {code}; period from official observation fields"))
    CANDIDATES[f"INE raw {code}"] = rows[:50]
    if not points:
        raise RuntimeError(f"INE {code} returned no observations; raw rows saved to candidates")
    return dedupe(points)

def ine_operation_candidates(operation: str, include_terms: list[str], label: str) -> list[dict[str, str]]:
    urls = [
        f"https://servicios.ine.es/wstempus/js/ES/SERIES_OPERACION/{operation}",
        f"https://servicios.ine.es/wstempus/js/ES/OPERACION/{operation}",
    ]
    rows: list[dict[str, str]] = []
    for url in urls:
        try:
            payload = request("GET", url).json()
            stack = payload if isinstance(payload, list) else [payload]
            while stack:
                item = stack.pop()
                if isinstance(item, dict):
                    code = clean(item.get("COD") or item.get("Codigo") or item.get("Id") or item.get("id"))
                    name = clean(item.get("Nombre") or item.get("name") or item.get("Descripcion"))
                    if code and name and all(term in norm(name) for term in include_terms):
                        rows.append({"code": code, "name": name, "discovery_url": url})
                    stack.extend(item.values())
                elif isinstance(item, list):
                    stack.extend(item)
        except Exception as exc:
            rows.append({"error": f"{url}: {type(exc).__name__}: {exc}"})
    CANDIDATES[label] = rows[:200]
    return [row for row in rows if row.get("code")]


def eurostat(dataset: str, filters: dict[str, str]) -> list[Point]:
    url = f"https://ec.europa.eu/eurostat/api/dissemination/statistics/1.0/data/{dataset}"
    response = request("GET", url, params={"format": "JSON", "lang": "EN", **filters})
    data = response.json()
    ids, sizes = data.get("id", []), data.get("size", [])
    if "time" not in ids or not data.get("value"):
        raise RuntimeError(f"Eurostat {dataset} returned no values")
    categories: dict[str, list[str]] = {}
    for dim in ids:
        index = data["dimension"][dim]["category"]["index"]
        if isinstance(index, dict):
            ordered = [""] * len(index)
            for code, pos in index.items():
                ordered[int(pos)] = code
            categories[dim] = ordered
        else:
            categories[dim] = list(index)
    points: list[Point] = []
    for raw_position, raw_value in data["value"].items():
        position = int(raw_position)
        coords = [0] * len(sizes)
        for i in range(len(sizes) - 1, -1, -1):
            coords[i] = position % sizes[i]
            position //= sizes[i]
        row = {dim: categories[dim][coords[i]] for i, dim in enumerate(ids)}
        period = period_key(row.get("time"))
        value = number(raw_value)
        if period and value is not None:
            points.append(Point(period, value, response.url, clean(data.get("status", {}).get(raw_position))))
    if not points:
        raise RuntimeError(f"Eurostat {dataset} parsed zero observations")
    return dedupe(points)


def parse_delimited(content: bytes, url: str) -> list[list[str]]:
    """Decode official European CSV files without destroying German labels."""
    text = None
    for encoding in ("utf-8-sig", "cp1252", "iso-8859-1"):
        try:
            candidate = content.decode(encoding)
            if "\ufffd" not in candidate:
                text = candidate
                break
        except UnicodeDecodeError:
            pass
    if text is None:
        text = content.decode("cp1252", errors="replace")
    sample = text[:10000]
    delimiter = ";" if sample.count(";") > sample.count(",") else ","
    return list(csv.reader(io.StringIO(text), delimiter=delimiter))


def destatis_table_csv(table: str) -> tuple[list[list[str]], str]:
    urls = [
        f"https://www-genesis.destatis.de/genesisWS/downloads/00/tables/{table}_00.csv",
        f"https://genesis.destatis.de/genesisWS/downloads/00/tables/{table}_00.csv",
    ]
    errors = []
    for url in urls:
        try:
            response = request("GET", url)
            return parse_delimited(response.content, response.url), response.url
        except Exception as exc:
            errors.append(f"{url}: {exc}")
    raise RuntimeError("Destatis CSV unavailable: " + " | ".join(errors))


def find_excel_links(page_url: str, required: list[str]) -> list[str]:
    response = request("GET", page_url)
    soup = BeautifulSoup(response.text, "html.parser")
    ranked: list[tuple[int, str]] = []
    for anchor in soup.find_all("a", href=True):
        href = urljoin(response.url, anchor.get("href", ""))
        context = norm(f"{anchor.get_text(' ', strip=True)} {anchor.parent.get_text(' ', strip=True) if anchor.parent else ''} {href}")
        suffix = href.lower().split("?")[0]
        downloadable = suffix.endswith((".xlsx", ".xls")) or "download" in href.lower() or "fileadmin" in href.lower()
        if not downloadable:
            continue
        score = sum(5 for term in required if term in context)
        score += 3 if suffix.endswith((".xlsx", ".xls")) else 0
        ranked.append((score, href))
    ranked.sort(reverse=True)
    return list(dict.fromkeys(url for _, url in ranked))

def workbook_cells(content: bytes, source_url: str = "") -> list[dict[str, Any]]:
    """Read both modern XLSX and legacy binary XLS workbooks."""
    is_xls = source_url.lower().split("?")[0].endswith(".xls") or content[:8] == bytes.fromhex("D0CF11E0A1B11AE1")
    out: list[dict[str, Any]] = []
    if is_xls:
        import xlrd
        book = xlrd.open_workbook(file_contents=content)
        for sheet in book.sheets():
            rows=[]
            for r in range(min(sheet.nrows,1000)):
                values=[]
                for c in range(sheet.ncols):
                    cell=sheet.cell(r,c)
                    value=cell.value
                    if cell.ctype == xlrd.XL_CELL_DATE:
                        value=xlrd.xldate_as_datetime(value,book.datemode).isoformat()
                    values.append(value)
                rows.append(values)
            out.append({"sheet": sheet.name, "rows": rows})
        return out
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    for ws in wb.worksheets:
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True), 1):
            rows.append(list(row))
            if i >= 1000:
                break
        out.append({"sheet": ws.title, "rows": rows})
    return out

def scan_workbook_for_period_values(content: bytes, source_url: str, expected: dict[str, float], keywords: list[str]) -> list[Point]:
    books = workbook_cells(content, source_url)
    CANDIDATES[f"Workbook preview {source_url}"] = [
        {"sheet":book["sheet"],"rows":book["rows"][:80]} for book in books
    ]
    candidates: list[dict[str, Any]] = []
    for sheet in books:
        sheet_name = norm(sheet["sheet"])
        rows = sheet["rows"]
        for r_idx, row in enumerate(rows):
            row_text = norm(" ".join(clean(x) for x in row if x is not None))
            for c_idx, cell in enumerate(row):
                period = period_key(cell)
                if not period:
                    text=clean(cell)
                    m=re.fullmatch(r"(0?[1-9]|1[0-2])[./-](\d{2})",text)
                    if m:
                        period=f"20{m.group(2)}-{int(m.group(1)):02d}"
                    else:
                        m=re.fullmatch(r"([A-Za-z]{3,9})[ -']?(\d{2})",text)
                        if m:
                            period=period_key(f"{m.group(1)} 20{m.group(2)}")
                if not period or period not in expected:
                    continue
                for rr in range(max(0, r_idx - 3), min(len(rows), r_idx + 4)):
                    for cc in range(max(0, c_idx - 3), min(len(rows[rr]), c_idx + 5)):
                        value = number(rows[rr][cc])
                        if value is None:
                            continue
                        context = norm(f"{sheet_name} {row_text} " + " ".join(clean(x) for x in rows[max(0, rr-2)][max(0, cc-2):cc+3]))
                        score = sum(3 for word in keywords if word in context)
                        score += max(0, 5 - abs(value - expected[period]) * 10)
                        candidates.append({"period": period, "value": value, "sheet": sheet["sheet"],
                                           "row": rr + 1, "column": cc + 1, "context": context[:400], "score": score})
    CANDIDATES[source_url] = sorted(candidates, key=lambda x: x["score"], reverse=True)[:100]
    points = []
    for period in expected:
        rows = [x for x in candidates if x["period"] == period]
        if rows:
            best = max(rows, key=lambda x: x["score"])
            points.append(Point(period, best["value"], source_url,
                                note=f"sheet={best['sheet']} row={best['row']} col={best['column']}"))
    if not points:
        raise RuntimeError("No expected periods found in workbook")
    return dedupe(points)


# ---------- Indicator-specific structured tests ----------

def spain_core_cpi() -> list[Point]:
    return ine_series("IPC208611")


def spain_unemployment() -> list[Point]:
    return ine_series("EPA815")


def spain_retail() -> list[Point]:
    # Discover series from ICM tables first; Tempus3 names vary by base year.
    discovery_urls = [
        "https://servicios.ine.es/wstempus/js/ES/TABLAS_OPERACION/ICM",
        "https://servicios.ine.es/wstempus/js/ES/SERIES_OPERACION/ICM",
    ]
    codes: dict[str,str] = {}
    errors = []
    for url in discovery_urls:
        try:
            payload = request("GET", url).json()
            stack = payload if isinstance(payload,list) else [payload]
            while stack:
                item=stack.pop()
                if isinstance(item,dict):
                    code=clean(item.get("COD") or item.get("Codigo") or item.get("Id") or item.get("id"))
                    name=clean(item.get("Nombre") or item.get("name") or item.get("Descripcion"))
                    if re.match(r"^[A-Za-z]+\d+$",code):
                        codes[code]=name
                    stack.extend(item.values())
                elif isinstance(item,list): stack.extend(item)
        except Exception as exc:
            errors.append(f"{url}: {exc}")
    ranked=[]
    for code,name in codes.items():
        n=norm(name)
        score=sum(2 for term in ["general","precios constantes","variacion anual"] if term in n)
        score += 8 if "total nacional" in n else -5
        score += 6 if "comercio al por menor excepto de vehiculos" in n else 0
        score -= 8 if "base 2005" in n else 0
        score -= 6 if "ajustados de estacionalidad" in n or "ajustados de calendario" in n else 0
        ranked.append((score,code,name))
    ranked.sort(reverse=True)
    CANDIDATES["西 零售 INE candidates"]=[{"score":a,"code":b,"name":c} for a,b,c in ranked[:200]]
    for _,code,name in ranked[:80]:
        try:
            points=ine_series(code)
            shared=[p for p in points if p.period in EXPECTED["西 零售"]]
            if shared and all(abs(p.value-EXPECTED["西 零售"][p.period]) <= 0.21 for p in shared):
                for p in points: p.note=f"INE discovered series {code}: {name}"
                return points
        except Exception as exc:
            errors.append(f"{code}: {exc}")
    raise RuntimeError("No matching INE ICM series found; "+" | ".join(errors[:8]))

def euro_core_cpi() -> list[Point]:
    errors=[]
    codes=["TOT_X_NRG_FOOD_NALC_TBC","TOT_X_NRG_FOOD"]
    for geo in ("EA21","EA20"):
        for dataset in ("prc_hicp_manr","prc_hicp_midx"):
            for coicop in codes:
                filters={"geo":geo,"coicop":coicop,
                         "sinceTimePeriod":"2025-01","untilTimePeriod":"2026-07"}
                filters["unit"]="RCH_A" if dataset=="prc_hicp_manr" else "I15"
                try:
                    points=eurostat(dataset,filters)
                    if dataset=="prc_hicp_midx":
                        points=yoy_from_levels({p.period:p.value for p in points},EXPECTED["歐 Core CPI"],points[-1].source_url)
                    if any(p.period in EXPECTED["歐 Core CPI"] and abs(p.value-EXPECTED["歐 Core CPI"][p.period])<=0.11 for p in points):
                        return points
                except Exception as exc:
                    errors.append(f"{dataset}/{geo}/{coicop}: {exc}")
    # Metadata discovery with only geo/time restrictions, retained in diagnostics.
    for geo in ("EA21","EA20"):
        try:
            url="https://ec.europa.eu/eurostat/api/dissemination/statistics/1.0/data/prc_hicp_manr"
            response=request("GET",url,params={"format":"JSON","lang":"EN","geo":geo,
                                               "sinceTimePeriod":"2026-04","untilTimePeriod":"2026-06"})
            payload=response.json()
            CANDIDATES[f"Eurostat core discovery {geo}"]={
                "id":payload.get("id"),"size":payload.get("size"),
                "dimensions":{k:list(v.get("category",{}).get("index",{}).keys())[:300]
                              for k,v in payload.get("dimension",{}).items()}
            }
        except Exception as exc:
            errors.append(f"discovery/{geo}: {exc}")
    raise RuntimeError("Eurostat core HICP candidates failed; "+" | ".join(errors))

def euro_unemployment() -> list[Point]:
    return eurostat("une_rt_m", {"geo": "EA21", "age": "TOTAL", "sex": "T", "unit": "PC_ACT", "s_adj": "SA"})


def insee_core_cpi() -> list[Point]:
    """Test structured INSEE series without querying the full BDM universe."""
    ids=["001768593"]  # official former annual underlying-inflation series
    errors=[]
    for idbank in ids:
        try:
            response=request("GET",f"https://api.insee.fr/series/BDM/V1/data/SERIES_BDM/{idbank}",
                             params={"lastNObservations":30})
            root=ET.fromstring(response.content); points=[]
            for obs in root.iter():
                if obs.tag.endswith("Obs"):
                    period=period_key(obs.attrib.get("TIME_PERIOD") or obs.attrib.get("timePeriod"))
                    value=number(obs.attrib.get("OBS_VALUE") or obs.attrib.get("obsValue"))
                    if period and value is not None:
                        points.append(Point(period,value,response.url,clean(obs.attrib.get("OBS_STATUS"))))
            CANDIDATES[f"INSEE idbank {idbank}"]=[asdict(p) for p in dedupe(points)]
            if any(p.period in EXPECTED["法 Core CPI"] for p in points):
                return dedupe(points)
        except Exception as exc:
            errors.append(f"{idbank}: {exc}")
    # Current 2026 series is not yet confirmed; preserve the official machine-readable data links page.
    raise RuntimeError("INSEE current-base core CPI idbank not confirmed; "+" | ".join(errors))

GERMAN_MONTHS={"januar":1,"februar":2,"marz":3,"april":4,"mai":5,"juni":6,"juli":7,"august":8,"september":9,"oktober":10,"november":11,"dezember":12}


def destatis_column_periods(rows: list[list[str]]) -> dict[int,str]:
    periods={}; years={}; width=max((len(r) for r in rows[:15]),default=0)
    for row in rows[:15]:
        for col,cell in enumerate(row):
            text=norm(cell)
            if re.fullmatch(r"20\d{2}",text): years[col]=int(text)
    current=None
    for col in range(width):
        if col in years: current=years[col]
        elif current is not None: years[col]=current
    for row in rows[:15]:
        for col,cell in enumerate(row):
            text=norm(cell); year=years.get(col)
            if not year: continue
            if text in GERMAN_MONTHS: periods[col]=f"{year:04d}-{GERMAN_MONTHS[text]:02d}"
            q=re.fullmatch(r"([1-4]) quartal",text)
            if q: periods[col]=f"{year:04d}-Q{q.group(1)}"
    return periods


def destatis_wide_series(rows: list[list[str]],url: str,row_terms: list[str],required_context: list[str]|None=None) -> dict[str,float]:
    periods=destatis_column_periods(rows); required_context=required_context or []; candidates=[]
    if not periods: raise RuntimeError("Destatis period headers not found")
    for idx,row in enumerate(rows):
        text=norm(" ".join(clean(x) for x in row[:8]))
        if not all(term in text for term in row_terms): continue
        if required_context and not all(term in text for term in required_context): continue
        values={period:number(row[col]) for col,period in periods.items() if col<len(row) and number(row[col]) is not None}
        if values: candidates.append({"row":idx+1,"text":text,"values":values})
    CANDIDATES[f"parsed {url} {'/'.join(row_terms)}"]=candidates
    if not candidates: raise RuntimeError(f"Destatis target row not found: {row_terms}")
    return max(candidates,key=lambda x:len(x["values"]))["values"]


def yoy_from_levels(levels: dict[str,float],expected: dict[str,float],url: str) -> list[Point]:
    points=[]
    for period in expected:
        if "-Q" in period:
            year,q=period.split("-Q"); prior=f"{int(year)-1:04d}-Q{q}"
        else:
            year,month=period.split("-"); prior=f"{int(year)-1:04d}-{month}"
        if period in levels and prior in levels and levels[prior]!=0:
            points.append(Point(period,(levels[period]/levels[prior]-1)*100,url,note="YoY from official levels"))
    if not points: raise RuntimeError("No same-period prior-year levels")
    return dedupe(points)


def best_wide_level_row_by_expected(rows: list[list[str]],url: str,expected: dict[str,float]) -> dict[str,float]:
    periods=destatis_column_periods(rows); ranked=[]
    for idx,row in enumerate(rows):
        levels={period:number(row[col]) for col,period in periods.items() if col<len(row) and number(row[col]) is not None}
        if len(levels)<12: continue
        errors=[]; computed={}
        for period,want in expected.items():
            year,month=period.split("-"); prior=f"{int(year)-1:04d}-{month}"
            if period in levels and prior in levels and levels[prior]!=0:
                got=(levels[period]/levels[prior]-1)*100; computed[period]=got; errors.append(abs(got-want))
        if errors: ranked.append({"row":idx+1,"label":" | ".join(clean(x) for x in row[:8]),"mean_abs_error":sum(errors)/len(errors),"matched":len(errors),"computed":computed,"levels":levels})
    ranked.sort(key=lambda x:(-x["matched"],x["mean_abs_error"]))
    CANDIDATES[f"Destatis ranked rows {url}"]=ranked[:30]
    if not ranked or ranked[0]["matched"]<max(1,len(expected)-1) or ranked[0]["mean_abs_error"]>0.15:
        raise RuntimeError("No Destatis row reproduced expected YoY pattern")
    return ranked[0]["levels"]


def destatis_core_cpi() -> list[Point]:
    rows, url = destatis_table_csv("61111-0006")
    attempts = [
        ["gesamtindex", "ohne", "nahrungsmittel", "energie"],
        ["ohne", "nahrungsmittel", "energie"],
    ]
    errors = []
    for terms in attempts:
        try:
            levels = destatis_wide_series(rows, url, terms)
            return yoy_from_levels(levels, EXPECTED["德 Core CPI"], url)
        except Exception as exc:
            errors.append(str(exc))
    try:
        levels = best_wide_level_row_by_expected(rows, url, EXPECTED["德 Core CPI"])
        return yoy_from_levels(levels, EXPECTED["德 Core CPI"], url)
    except Exception as exc:
        errors.append(str(exc))
    CANDIDATES["Destatis raw 61111-0006"] = rows[:250]
    raise RuntimeError("Destatis core CPI row unresolved; " + " | ".join(errors))


def destatis_industry() -> list[Point]:
    rows, url = destatis_table_csv("42153-0001")
    # Calendar-adjusted level. Use total industry/Produzierendes Gewerbe, then YoY.
    attempts = [
        (["produzierendes", "gewerbe"], ["kalenderbereinigt"]),
        (["produzierendes", "gewerbe"], []),
    ]
    errors = []
    for terms, context in attempts:
        try:
            levels = destatis_wide_series(rows, url, terms, context)
            return yoy_from_levels(levels, EXPECTED["德 工業"], url)
        except Exception as exc:
            errors.append(str(exc))
    try:
        levels = best_wide_level_row_by_expected(rows, url, EXPECTED["德 工業"])
        return yoy_from_levels(levels, EXPECTED["德 工業"], url)
    except Exception as exc:
        errors.append(str(exc))
    CANDIDATES["Destatis raw 42153-0001"] = rows[:250]
    raise RuntimeError("Destatis industry row unresolved; " + " | ".join(errors))


def destatis_gdp() -> list[Point]:
    """Parse the confirmed quarterly table 81000-0002."""
    rows, url = destatis_table_csv("81000-0002")
    periods = destatis_column_periods(rows)
    target = None
    inspected = []
    for idx, row in enumerate(rows):
        row_text = norm(" ".join(clean(x) for x in row[:6]))
        if all(term in row_text for term in ["originalwerte", "preisbereinigt", "bruttoinlandsprodukt", "veranderung", "prozent"]):
            values = {period: number(row[col]) for col, period in periods.items() if col < len(row) and number(row[col]) is not None}
            inspected.append({"row": idx + 1, "text": row_text, "values": values})
            if values:
                target = values
                break
    CANDIDATES["德 GDP 81000-0002 parsed"] = inspected
    if not target:
        raise RuntimeError("Confirmed GDP percentage row not found in 81000-0002")
    return [Point(period, target[period], url, note="Originalwerte; price-adjusted GDP YoY")
            for period in EXPECTED["德 GDP"] if period in target]

def ifo_excel() -> list[Point]:
    pages = [
        "https://www.ifo.de/umfragen/zeitreihen",
        "https://www.ifo.de/en/umfragen/time-series",
        "https://www.ifo.de/ifo-zeitreihen",
        "https://www.ifo.de/en/ifo-time-series",
    ]
    links: list[str] = []
    for page in pages:
        try:
            links.extend(find_excel_links(page, ["geschaftsklima"] if "/ifo-zeitreihen" in page else ["business", "climate"]))
        except Exception as exc:
            CANDIDATES.setdefault("ifo link errors", []).append(f"{page}: {exc}")
    links = list(dict.fromkeys(links))
    CANDIDATES["ifo Excel links"] = links
    errors = []
    for link in links[:20]:
        try:
            response = request("GET", link)
            return scan_workbook_for_period_values(response.content, response.url, EXPECTED["德 企業信心"],
                                                   ["deutschland", "geschaftsklima", "index", "saisonbereinigt"])
        except Exception as exc:
            errors.append(f"{link}: {exc}")
    raise RuntimeError("ifo official XLSX not parsed; " + " | ".join(errors[:5]))


def zew_excel(which: str) -> list[Point]:
    page = "https://www.zew.de/en/publications/zew-expertises-research-reports/research-reports/business-cycle/zew-financial-market-survey"
    links = find_excel_links(page, ["historical", "time", "series"])
    CANDIDATES["ZEW Excel links"] = links
    expected = EXPECTED["德信心 Current" if which == "current" else "德信心 expect"]
    keywords = ["economic", "situation", "germany"] if which == "current" else ["economic", "sentiment", "germany"]
    errors = []
    for link in links[:20]:
        try:
            response = request("GET", link)
            points = scan_workbook_for_period_values(response.content, response.url, expected, keywords)
            if points:
                return points
        except Exception as exc:
            errors.append(f"{link}: {exc}")
    raise RuntimeError(f"ZEW {which} official Excel not parsed; " + " | ".join(errors[:5]))


def tgss_pxweb() -> list[Point]:
    roots = [
        "https://w6.seg-social.es/PXWeb/api/v1/es",
        "https://w6.seg-social.es/PXWeb/api/v1/es/Afiliados%20en%20alta%20laboral",
    ]
    tables: list[str] = []
    visited: set[str] = set()
    errors = []

    def walk(url: str, depth: int = 0) -> None:
        if url in visited or depth > 5 or len(visited) > 200:
            return
        visited.add(url)
        try:
            payload = request("GET", url).json()
            if not isinstance(payload, list):
                return
            for item in payload:
                item_id = clean(item.get("id"))
                item_type = clean(item.get("type")).lower()
                child = f"{url.rstrip('/')}/{requests.utils.quote(item_id, safe='')}"
                if item_type in {"l", "folder", "directory"}:
                    walk(child, depth + 1)
                elif item_type in {"t", "table"} or item_id.lower().endswith(".px"):
                    tables.append(child)
        except Exception as exc:
            errors.append(f"{url}: {exc}")

    for root in roots:
        walk(root)
    CANDIDATES["TGSS PXWeb tables"] = tables
    ranked = []
    for table in tables:
        try:
            meta = request("GET", table).json()
            text = norm(json.dumps(meta, ensure_ascii=False))
            score = sum(3 for term in ["desestacional", "ajuste estacional", "total sistema", "variacion mensual"] if term in text)
            if score:
                ranked.append((score, table, meta))
        except Exception as exc:
            errors.append(f"{table}: {exc}")
    ranked.sort(reverse=True, key=lambda x: x[0])
    CANDIDATES["TGSS ranked adjusted tables"] = [{"score": s, "url": u, "meta": m} for s, u, m in ranked[:20]]
    # Query only tables whose metadata explicitly supports adjusted total/month.
    for _, table, meta in ranked[:10]:
        try:
            query = []
            for variable in meta.get("variables", []):
                code = variable["code"]
                values = variable.get("values", [])
                texts = variable.get("valueTexts", [])
                selected = []
                for value, text in zip(values, texts):
                    n = norm(text)
                    if any(term in n for term in ["2026", "2025", "total", "desestacional", "ajustada"]):
                        selected.append(value)
                query.append({"code": code, "selection": {"filter": "item", "values": selected[:40] or values[-20:]}})
            response = request("POST", table, json={"query": query, "response": {"format": "json-stat2"}})
            data = response.json()
            CANDIDATES[f"TGSS response {table}"] = data
            # Generic json-stat2 parse through a temporary Eurostat-like adapter.
            ids = data.get("id", []); sizes = data.get("size", [])
            dimensions = data.get("dimension", {})
            values = data.get("value", [])
            if not ids or not sizes or not values:
                continue
            categories = {}
            for dim in ids:
                index = dimensions[dim]["category"]["index"]
                if isinstance(index, dict):
                    ordered = [""] * len(index)
                    for code, pos in index.items(): ordered[int(pos)] = code
                    categories[dim] = ordered
                else: categories[dim] = list(index)
            points = []
            iterable = enumerate(values) if isinstance(values, list) else ((int(k), v) for k, v in values.items())
            for position, raw_value in iterable:
                if raw_value is None: continue
                coords = [0] * len(sizes); pos = position
                for i in range(len(sizes)-1, -1, -1): coords[i] = pos % sizes[i]; pos //= sizes[i]
                row = {dim: categories[dim][coords[i]] for i, dim in enumerate(ids)}
                period = next((period_key(v) for v in row.values() if period_key(v)), None)
                value = number(raw_value)
                if period and value is not None:
                    # Values may be persons; convert only when clearly too large for thousand-person EU_ECON data.
                    if abs(value) > 10000: value /= 1000.0
                    points.append(Point(period, value, response.url, note="TGSS PXWeb adjusted-series candidate"))
            if points:
                return dedupe(points)
        except Exception as exc:
            errors.append(f"{table}: {exc}")
    raise RuntimeError("TGSS PXWeb adjusted monthly series not confirmed; " + " | ".join(errors[:8]))


TESTS: list[SourceTest] = [
    SourceTest("西Core CPI", "INE Tempus3 IPC208611", "JSON API", True,
               "National CPI excluding unprocessed food and energy, YoY", spain_core_cpi),
    SourceTest("法 Core CPI", "INSEE BDM/Melodi structured discovery", "XML/JSON API", True,
               "National underlying/core inflation, YoY", insee_core_cpi),
    SourceTest("德 Core CPI", "Destatis 61111-0006", "CSV", True,
               "CPI excluding food and energy; YoY from official index levels", destatis_core_cpi),
    SourceTest("歐 Core CPI", "Eurostat HICP", "JSON API", True,
               "HICP excluding energy, food, alcohol and tobacco, YoY", euro_core_cpi),
    SourceTest("西 失業率", "INE Tempus3 EPA815", "JSON API", True,
               "EPA national unemployment rate, quarterly", spain_unemployment),
    SourceTest("歐 失業率", "Eurostat une_rt_m", "JSON API", True,
               "ILO unemployment rate, monthly SA", euro_unemployment),
    SourceTest("西 就業", "TGSS PXWeb", "PXWeb JSON API/CSV/XLSX", True,
               "Total-system seasonally adjusted affiliation monthly change, thousand", tgss_pxweb),
    SourceTest("西 零售", "INE Tempus3 ICM discovered series", "JSON API", True,
               "Original constant-price retail index, YoY", spain_retail),
    SourceTest("德 工業", "Destatis 42153-0001", "CSV", True,
               "Real production in industry, calendar-adjusted YoY", destatis_industry),
    SourceTest("德信心 Current", "ZEW historical time series", "XLSX", True,
               "Economic Situation Germany, balance", lambda: zew_excel("current")),
    SourceTest("德信心 expect", "ZEW historical time series", "XLSX", True,
               "ZEW Indicator of Economic Sentiment Germany, balance", lambda: zew_excel("expect")),
    SourceTest("德 企業信心", "ifo Business Climate time series", "XLSX", True,
               "Germany index 2015=100, seasonally adjusted", ifo_excel),
    SourceTest("德 GDP", "Destatis 81000 national accounts", "CSV", True,
               "Price-adjusted GDP YoY, not calendar adjusted", destatis_gdp),
]


def compare(points: list[Point], expected: dict[str, float], tolerance: float) -> dict[str, Any]:
    mapping = {point.period: point for point in points}
    rows = []
    for period, expected_value in sorted(expected.items(), reverse=True):
        point = mapping.get(period)
        if point is None:
            rows.append({"period": period, "EU_ECON": expected_value, "official": None,
                         "difference": None, "match": False, "note": "official period absent"})
            continue
        difference = point.value - expected_value
        rows.append({"period": period, "EU_ECON": expected_value, "official": point.value,
                     "difference": difference, "match": abs(difference) <= tolerance,
                     "source_url": point.source_url, "status": point.status, "note": point.note})
    matched = [row for row in rows if row["official"] is not None]
    if matched and all(row["match"] for row in matched) and len(matched) == len(rows):
        status = "MATCH_ALL"
    elif matched and all(row["match"] for row in matched):
        status = "MATCH_AVAILABLE_PERIODS"
    elif matched:
        status = "VALUE_MISMATCH"
    else:
        status = "NO_SAME_PERIOD"
    return {"status": status, "rows": rows, "matched_periods": len(matched), "expected_periods": len(rows)}


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--out", default="debug/eu_structured_sources")
    parser.add_argument("--tolerance", type=float, default=0.051)
    args = parser.parse_args()
    out = Path(args.out)
    out.mkdir(parents=True, exist_ok=True)
    report = {"script_version": VERSION, "generated_at": datetime.now(timezone.utc).isoformat(),
              "policy": "official API/JSON/SDMX/CSV/XLSX/PX-Web only; PMI excluded",
              "tolerance": args.tolerance, "results": []}
    summary: dict[str, int] = {}
    for test in TESTS:
        log(f"\n[TEST] {test.label} | {test.source_name} | {test.format}")
        item = {"label": test.label, "source_name": test.source_name, "format": test.format,
                "official": test.official, "definition": test.definition}
        try:
            points = dedupe(test.fetcher())
            item["latest_points"] = [asdict(p) for p in points[-24:]]
            item["comparison"] = compare(points, EXPECTED[test.label], args.tolerance)
            item["status"] = item["comparison"]["status"]
            log(f"[RESULT] {item['status']} | points={len(points)}")
        except Exception as exc:
            item["status"] = "FETCH_ERROR"
            item["error"] = f"{type(exc).__name__}: {exc}"
            item["traceback"] = traceback.format_exc(limit=8)
            log(f"[ERROR] {item['error']}")
        summary[item["status"]] = summary.get(item["status"], 0) + 1
        report["results"].append(item)
    report["summary"] = summary

    (out / "structured_source_comparison.json").write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    (out / "structured_source_candidates.json").write_text(json.dumps(CANDIDATES, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    (out / "structured_source_http_log.json").write_text(json.dumps(HTTP_LOG, ensure_ascii=False, indent=2), encoding="utf-8")

    with (out / "structured_source_comparison.csv").open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(["label", "status", "source_name", "format", "definition", "period",
                         "EU_ECON", "official", "difference", "match", "source_url", "note_or_error"])
        for item in report["results"]:
            comparison = item.get("comparison")
            if not comparison:
                writer.writerow([item["label"], item["status"], item["source_name"], item["format"],
                                 item["definition"], "", "", "", "", "", "", item.get("error", "")])
                continue
            for row in comparison["rows"]:
                writer.writerow([item["label"], item["status"], item["source_name"], item["format"],
                                 item["definition"], row.get("period"), row.get("EU_ECON"), row.get("official"),
                                 row.get("difference"), row.get("match"), row.get("source_url", ""), row.get("note", "")])

    log("\n=== SUMMARY ===")
    for key, value in summary.items():
        log(f"{key}: {value}")
    log(f"Outputs: {out}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
