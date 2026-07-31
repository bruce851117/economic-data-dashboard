#!/usr/bin/env python3
"""Validate Australian macro sources against the user's AU_ECON reference values.

Source policy:
1. Official API / CSV / XLSX
2. Official PDF
3. Official HTML release only when no public structured series is available

Read-only diagnostic. It never modifies production JSON files.
"""
from __future__ import annotations

import csv
import html
import io
import json
import math
import re
import sys
import traceback
from dataclasses import asdict, dataclass
from datetime import date, datetime
from datetime import timedelta as dt_timedelta
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any
from urllib.parse import urljoin

import requests
from bs4 import BeautifulSoup, NavigableString
from openpyxl import load_workbook
from pypdf import PdfReader

VERSION = "2026-07-31-au-source-validation-v15-exact-series-flexible-xls"
OUT = Path("debug/au_macro_sources")
ABS_API = "https://data.api.abs.gov.au/rest/data"
SESSION = requests.Session()
SESSION.headers.update({
    "User-Agent": "Mozilla/5.0 (compatible; AUMacroSourceValidator/1.0; GitHub-Actions)",
    "Accept-Language": "en-AU,en;q=0.9",
})


@dataclass
class Point:
    period: str
    value: float
    source_url: str
    status: str = ""
    note: str = ""


@dataclass
class Target:
    label: str
    frequency: str
    expected: dict[str, float]
    source_tier: str
    source_name: str


def log(message: str) -> None:
    print(message, flush=True)


def clean(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def norm(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", " ", clean(value).lower()).strip()


def number(value: Any) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        value = float(value)
        return value if math.isfinite(value) else None
    text = clean(value).replace("%", "").replace(",", "").replace("−", "-")
    match = re.search(r"[+-]?\d+(?:\.\d+)?", text)
    return float(match.group()) if match else None


def period_key(value: Any) -> str | None:
    # openpyxl returns ABS workbook date cells as datetime/date objects.
    if isinstance(value, (datetime, date)):
        month = value.month
        quarter = (month - 1) // 3 + 1
        # National accounts workbooks use Mar/Jun/Sep/Dec quarter dates.
        if month in {3, 6, 9, 12}:
            return f"{value.year:04d}-Q{quarter}"
        return f"{value.year:04d}-{month:02d}"
    text = clean(value)
    # Also accept displayed Excel timestamps and historical dates.
    timestamp = re.fullmatch(r"((?:19|20)\d{2})[-/](0?[1-9]|1[0-2])[-/](0?[1-9]|[12]\d|3[01])(?:[ T].*)?", text)
    if timestamp:
        year, month = int(timestamp.group(1)), int(timestamp.group(2))
        if month in {3, 6, 9, 12}:
            return f"{year:04d}-Q{(month - 1) // 3 + 1}"
        return f"{year:04d}-{month:02d}"
    monthly = re.fullmatch(r"((?:19|20)\d{2})[-/]?(0[1-9]|1[0-2])", text)
    if monthly:
        return f"{monthly.group(1)}-{monthly.group(2)}"
    quarterly = re.fullmatch(r"((?:19|20)\d{2})[- ]?[Qq]([1-4])", text)
    if quarterly:
        return f"{quarterly.group(1)}-Q{quarterly.group(2)}"
    named_quarter = re.fullmatch(
        r"(Mar|Jun|Sep|Dec)(?:ember)?(?:[- ]+Qtr)?[- ]+((?:19|20)\d{2})",
        text,
        re.I,
    )
    if named_quarter:
        quarter_map = {"mar": 1, "jun": 2, "sep": 3, "dec": 4}
        return f"{int(named_quarter.group(2)):04d}-Q{quarter_map[named_quarter.group(1)[:3].lower()]}"
    # ABS labels such as Jun-2026 or May-26.
    month_map = {name.lower(): index for index, name in enumerate(
        "Jan Feb Mar Apr May Jun Jul Aug Sep Oct Nov Dec".split(), 1)}
    named = re.fullmatch(r"([A-Za-z]{3})[- ]((?:19|20)\d{2}|\d{2})", text)
    if named and named.group(1).lower() in month_map:
        year = int(named.group(2))
        if year < 100:
            year += 2000
        return f"{year:04d}-{month_map[named.group(1).lower()]:02d}"
    return None


def get(url: str, **kwargs: Any) -> requests.Response:
    response = SESSION.get(url, timeout=90, allow_redirects=True, **kwargs)
    log(f"[HTTP] {response.status_code} {response.url} bytes={len(response.content)}")
    response.raise_for_status()
    return response


def dedupe(points: list[Point]) -> list[Point]:
    mapping = {point.period: point for point in points}
    return [mapping[key] for key in sorted(mapping)]


ABS_FLOW_ALIASES = {
    "LF": ["LF"],
    "JV": ["JV"],
    "WPI": ["WPI"],
    # Monthly CPI moved into the consolidated CPI flow. CPI_M is retained as a
    # fallback for historical metadata snapshots.
    "CPI_MONTHLY": ["CPI", "CPI_M"],
    "HSI_M": ["HSI_M", "MHSI", "HSI"],
    "ANA_AGG": ["ANA_AGG"],
}


def abs_csv(flow: str, start_period: str, *, last_n: int | None = None) -> tuple[list[dict[str, str]], str]:
    """Download ABS SDMX CSV using the official media-type labels syntax.

    ABS expects labels=both inside the Accept media type, not as a URL query
    parameter. The v1 diagnostic incorrectly sent labels=both in the query,
    which caused the common HTTP 400 responses.
    """
    params: dict[str, str] = {
        "startPeriod": start_period,
        "dimensionAtObservation": "TIME_PERIOD",
    }
    if last_n:
        params["lastNObservations"] = str(last_n)
    errors: list[str] = []
    for candidate in ABS_FLOW_ALIASES.get(flow, [flow]):
        url = f"{ABS_API}/ABS,{candidate}/all"
        response = SESSION.get(
            url,
            params=params,
            timeout=120,
            allow_redirects=True,
            headers={"Accept": "application/vnd.sdmx.data+csv;labels=both"},
        )
        log(f"[HTTP] {response.status_code} {response.url} bytes={len(response.content)}")
        safe_name = re.sub(r"[^a-z0-9]+", "_", candidate.lower()).strip("_")
        (OUT / f"abs_{safe_name}_http_{response.status_code}.txt").write_bytes(response.content)
        if response.status_code >= 400:
            errors.append(f"{candidate}: HTTP {response.status_code}: {response.text[:500]}")
            continue
        raw_path = OUT / f"abs_{safe_name}_raw.csv"
        raw_path.write_bytes(response.content)
        text = response.content.decode("utf-8-sig", "replace")
        reader = csv.DictReader(io.StringIO(text))
        raw_rows = list(reader)
        (OUT / f"abs_{safe_name}_columns.json").write_text(
            json.dumps({"columns": reader.fieldnames or []}, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        rows = []
        for raw_row in raw_rows:
            row = {str(key).strip(): value for key, value in raw_row.items() if key is not None}
            # ABS SDMX-CSV responses may label observation columns differently
            # depending on representation. Provide canonical aliases.
            for key, value in list(row.items()):
                # labels=both produces headers such as
                # TIME_PERIOD:Time period and OBS_VALUE:Observation value.
                component_id = key.split(":", 1)[0].strip()
                compact = re.sub(r"[^A-Z0-9]", "", component_id.upper())
                if compact in {"TIMEPERIOD", "TIME", "PERIOD"} and not row.get("TIME_PERIOD"):
                    row["TIME_PERIOD"] = value
                if compact in {"OBSVALUE", "VALUE", "OBSERVATIONVALUE"} and not row.get("OBS_VALUE"):
                    row["OBS_VALUE"] = value
            rows.append(row)
        if rows:
            return rows, response.url
        errors.append(f"{candidate}: successful response contained no CSV rows")
    raise RuntimeError("ABS dataflow attempts failed: " + " | ".join(errors))

def component_id(key: str) -> str:
    return re.sub(r"[^A-Z0-9_]", "", key.split(":", 1)[0].strip().upper())


def row_text(row: dict[str, Any]) -> str:
    ignored = {"TIME_PERIOD", "OBS_VALUE", "OBS_STATUS", "OBS_COMMENT", "UNIT_MULT", "DECIMALS"}
    return norm(" ".join(
        str(value) for key, value in row.items()
        if component_id(str(key)) not in ignored
    ))


def series_identity(row: dict[str, Any]) -> str:
    # labels=both retains labelled source columns. TIME_PERIOD must be excluded
    # by SDMX component ID, otherwise each observation becomes its own series.
    ignored = {"TIME_PERIOD", "OBS_VALUE", "OBS_STATUS", "OBS_COMMENT", "UNIT_MULT", "DECIMALS"}
    return "|".join(
        f"{key}={row.get(key, '')}"
        for key in sorted(row)
        if component_id(str(key)) not in ignored
    )


def rank_abs_series(
    rows: list[dict[str, str]],
    include: list[str],
    exclude: list[str],
    expected: dict[str, float],
    *,
    transform: str = "level",
) -> tuple[list[Point], list[dict[str, Any]]]:
    """Rank every ABS series by numerical agreement, then metadata relevance.

    ABS label columns vary across dataflows, so requiring every English phrase
    to appear in one concatenated label incorrectly discarded valid series.
    The diagnostic now keeps every numeric series, compares common reference
    periods first, and uses metadata terms only as a secondary ranking signal.
    """
    grouped: dict[str, list[dict[str, str]]] = {}
    for row in rows:
        grouped.setdefault(series_identity(row), []).append(row)

    ranked: list[dict[str, Any]] = []
    for identity, series_rows in grouped.items():
        text = row_text(series_rows[0])
        include_hits = sum(1 for term in include if term in text)
        exclude_hits = sum(1 for term in exclude if term in text)
        metadata_score = include_hits * 10 - exclude_hits * 20

        levels: dict[str, float] = {}
        for row in series_rows:
            period = period_key(row.get("TIME_PERIOD"))
            value = number(row.get("OBS_VALUE"))
            if period and value is not None:
                levels[period] = value
        if not levels:
            continue

        values = levels
        if transform in {"mom_diff", "mom_diff_thousands"}:
            ordered = sorted(levels)
            divisor = 1000.0 if transform == "mom_diff_thousands" else 1.0
            values = {
                current: (levels[current] - levels[previous]) / divisor
                for previous, current in zip(ordered, ordered[1:])
            }
        elif transform == "yoy_pct_q":
            ordered = sorted(levels)
            values = {
                ordered[index]: (levels[ordered[index]] / levels[ordered[index - 4]] - 1.0) * 100.0
                for index in range(4, len(ordered))
                if levels[ordered[index - 4]] != 0
            }

        common = sorted(set(expected) & set(values))
        diffs = [abs(values[p] - expected[p]) for p in common]
        mae = sum(diffs) / len(diffs) if diffs else 999999.0
        max_error = max(diffs) if diffs else 999999.0
        ranked.append({
            "identity": identity,
            "description": text,
            "include_hits": include_hits,
            "include_total": len(include),
            "exclude_hits": exclude_hits,
            "metadata_score": metadata_score,
            "matches": len(common),
            "mae": mae,
            "max_error": max_error,
            "latest": [[period, values[period]] for period in sorted(values)[-12:]],
            "values": values,
        })

    ranked.sort(
        key=lambda item: (
            -item["matches"],
            item["mae"],
            item["max_error"],
            -item["metadata_score"],
            -len(item["values"]),
        )
    )
    if not ranked:
        raise RuntimeError("ABS response contained no numeric time series")

    best = ranked[0]
    points = [
        Point(period, value, "ABS Data API", note=best["description"])
        for period, value in sorted(best["values"].items())
    ]
    public_candidates = [
        {key: value for key, value in item.items() if key != "values"}
        for item in ranked[:50]
    ]
    return points, public_candidates

def fetch_abs_target(flow: str, start: str, include: list[str], exclude: list[str], expected: dict[str, float], transform: str = "level") -> tuple[list[Point], dict[str, Any]]:
    rows, url = abs_csv(flow, start)
    points, candidates = rank_abs_series(rows, include, exclude, expected, transform=transform)
    if flow == "JV":
        quarter_month = {"Q1": "02", "Q2": "05", "Q3": "08", "Q4": "11"}
        for point in points:
            match = re.fullmatch(r"(\d{4})-(Q[1-4])", point.period)
            if match:
                point.period = f"{match.group(1)}-{quarter_month[match.group(2)]}"
    for point in points:
        point.source_url = url
    return points, {"flow": flow, "request_url": url, "candidates": candidates}


def workbook_rows(content: bytes) -> list[dict[str, Any]]:
    if content.startswith(b"PK"):
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        return [{"sheet": sheet.title, "rows": [list(row) for row in sheet.iter_rows(values_only=True)]} for sheet in workbook.worksheets]
    if content.startswith(bytes.fromhex("D0CF11E0")):
        import xlrd
        workbook = xlrd.open_workbook(file_contents=content)
        return [{"sheet": sheet.name, "rows": [sheet.row_values(row) for row in range(sheet.nrows)]} for sheet in workbook.sheets()]
    raise RuntimeError("Downloaded file is neither XLSX nor XLS")


def excel_period(value: Any, frequency: str = "Q") -> str | None:
    if isinstance(value, (int, float)) and 20000 <= float(value) <= 80000:
        converted = datetime(1899, 12, 30) + dt_timedelta(days=float(value))
        if frequency == "Q":
            return f"{converted.year:04d}-Q{(converted.month - 1) // 3 + 1}"
        return f"{converted.year:04d}-{converted.month:02d}"
    period = period_key(value)
    if period and frequency == "Q" and re.fullmatch(r"20\d{2}-\d{2}", period):
        year, month = map(int, period.split("-"))
        return f"{year:04d}-Q{(month - 1) // 3 + 1}"
    return period


def exact_series_from_workbook(content: bytes, series_id: str, source_url: str, frequency: str = "Q") -> tuple[list[Point], dict[str, Any]]:
    for book in workbook_rows(content):
        rows = book["rows"]
        for id_row, row in enumerate(rows):
            for value_col, cell in enumerate(row):
                if clean(cell).upper() != series_id.upper():
                    continue
                values: dict[str, float] = {}
                for data_row in rows[id_row + 1:]:
                    if not data_row:
                        continue
                    period = excel_period(data_row[0], frequency)
                    if period and value_col < len(data_row):
                        value = number(data_row[value_col])
                        if value is not None:
                            values[period] = value
                if not values:
                    raise RuntimeError(f"Series {series_id} found at {book['sheet']} row {id_row + 1}, but no dated values followed")
                points = [Point(period, value, source_url, note=f"sheet={book['sheet']}; series={series_id}") for period, value in sorted(values.items())]
                return points, {"sheet": book["sheet"], "series_id": series_id, "series_id_row": id_row + 1, "value_col": value_col + 1, "observation_count": len(values)}
    raise RuntimeError(f"Series ID {series_id} not found in downloaded workbook")


def fetch_anz_job_ads(expected: dict[str, float]) -> tuple[list[Point], dict[str, Any]]:
    url = "https://www.anz.com.au/content/dam/anzcomau/mediacentre/pdfs/jobads/2026/july/ANZ-Indeed%20Australian%20Job%20Ads%20data_Jun26.xlsx"
    response = get(url)
    (OUT / "anz_job_ads_raw.xlsx").write_bytes(response.content)
    books = workbook_rows(response.content)
    candidates: list[dict[str, Any]] = []
    for book in books:
        rows = book["rows"]
        for date_col in range(0, min(5, max((len(row) for row in rows), default=0))):
            periods: dict[int, str] = {}
            for index, row in enumerate(rows):
                if date_col >= len(row):
                    continue
                value = row[date_col]
                if hasattr(value, "year") and hasattr(value, "month"):
                    periods[index] = f"{value.year:04d}-{value.month:02d}"
                else:
                    period = period_key(value)
                    if period:
                        periods[index] = period
            if len(periods) < 10:
                continue
            max_cols = max((len(row) for row in rows), default=0)
            for value_col in range(max_cols):
                values: dict[str, float] = {}
                for index, period in periods.items():
                    if value_col < len(rows[index]):
                        value = number(rows[index][value_col])
                        if value is not None:
                            values[period] = value
                common = sorted(set(values) & set(expected))
                if not common:
                    continue
                mae = sum(abs(values[p] - expected[p]) for p in common) / len(common)
                candidates.append({"sheet": book["sheet"], "date_col": date_col + 1, "value_col": value_col + 1, "matches": len(common), "mae": mae, "values": values})
    candidates.sort(key=lambda item: (-item["matches"], item["mae"]))
    if not candidates:
        raise RuntimeError("ANZ XLSX contained no series matching reference periods")
    best = candidates[0]
    points = [Point(period, value, response.url, note=f"sheet={best['sheet']}; col={best['value_col']}") for period, value in sorted(best["values"].items())]
    return points, {"request_url": response.url, "candidates": [{k: v for k, v in item.items() if k != "values"} for item in candidates[:30]]}


def fetch_abs_workbook_candidate(
    urls: list[str],
    expected: dict[str, float],
    raw_name: str,
    note: str,
) -> tuple[list[Point], dict[str, Any]]:
    errors: list[str] = []
    for url in urls:
        try:
            response = get(url)
            (OUT / raw_name).write_bytes(response.content)
            books = workbook_rows(response.content)
            candidates: list[dict[str, Any]] = []
            for book in books:
                rows = book["rows"]
                max_cols = max((len(row) for row in rows), default=0)
                for date_col in range(max_cols):
                    dates: dict[int, str] = {}
                    for row_index, row in enumerate(rows):
                        if date_col >= len(row):
                            continue
                        period = period_key(row[date_col])
                        if period:
                            dates[row_index] = period
                    if not dates:
                        continue
                    for value_col in range(max_cols):
                        values: dict[str, float] = {}
                        for row_index, period in dates.items():
                            if value_col < len(rows[row_index]):
                                value = number(rows[row_index][value_col])
                                if value is not None:
                                    values[period] = value
                        # First compare a directly published YoY column.
                        common = sorted(set(values) & set(expected))
                        if common:
                            diffs = [abs(values[item] - expected[item]) for item in common]
                            candidates.append({
                                "sheet": book["sheet"], "date_col": date_col + 1,
                                "value_col": value_col + 1, "transform": "level",
                                "matches": len(common), "mae": sum(diffs) / len(diffs),
                                "values": values,
                            })
                        # Also test YoY calculated from quarterly level columns.
                        ordered = sorted(values)
                        yoy = {
                            ordered[index]: (values[ordered[index]] / values[ordered[index - 4]] - 1.0) * 100.0
                            for index in range(4, len(ordered))
                            if values[ordered[index - 4]] != 0
                        }
                        common_yoy = sorted(set(yoy) & set(expected))
                        if common_yoy:
                            diffs = [abs(yoy[item] - expected[item]) for item in common_yoy]
                            candidates.append({
                                "sheet": book["sheet"], "date_col": date_col + 1,
                                "value_col": value_col + 1, "transform": "yoy_pct_q",
                                "matches": len(common_yoy), "mae": sum(diffs) / len(diffs),
                                "values": yoy,
                            })
            candidates.sort(key=lambda item: (-item["matches"], item["mae"]))
            if not candidates:
                raise RuntimeError("workbook contained no quarterly candidate series")
            best = candidates[0]
            points = [
                Point(period, value, response.url, note=f"{note}; sheet={best['sheet']}; col={best['value_col']}; transform={best['transform']}")
                for period, value in sorted(best["values"].items())
            ]
            return points, {
                "request_url": response.url,
                "note": note,
                "candidates": [{key: value for key, value in item.items() if key != "values"} for item in candidates[:50]],
            }
        except Exception as error:
            errors.append(f"{url}: {type(error).__name__}: {error}")
    raise RuntimeError("ABS national accounts workbook attempts failed: " + " | ".join(errors))


def discover_abs_workbook(landing_urls: list[str], filename_stem: str, title_token: str, fallback_urls: list[str]) -> tuple[requests.Response, dict[str, Any]]:
    candidates: list[str] = []
    discovery: list[dict[str, Any]] = []
    for landing_url in landing_urls:
        try:
            page = get(landing_url)
            soup = BeautifulSoup(page.text, "html.parser")
            found: list[str] = []
            for anchor in soup.find_all("a", href=True):
                href = anchor.get("href", "")
                context = clean(" ".join([anchor.get_text(" ", strip=True), anchor.parent.get_text(" ", strip=True) if anchor.parent else ""]))
                filename_match = filename_stem.lower() in href.lower()
                title_match = title_token.lower() in context.lower()
                if (filename_match or title_match) and re.search(r"\.xlsx?$", href, re.I):
                    url = urljoin(page.url, href)
                    if url not in candidates:
                        candidates.append(url); found.append(url)
            discovery.append({"landing_url": page.url, "found": found})
        except Exception as error:
            discovery.append({"landing_url": landing_url, "error": f"{type(error).__name__}: {error}"})
    for url in fallback_urls:
        if url not in candidates:
            candidates.append(url)
    errors=[]
    for url in candidates:
        try:
            response=get(url)
            if response.content.startswith((b"PK", bytes.fromhex("D0CF11E0"))):
                return response, {"discovery": discovery, "candidates": candidates}
            errors.append(f"{url}: invalid workbook signature")
        except Exception as error:
            errors.append(f"{url}: {type(error).__name__}: {error}")
    raise RuntimeError("Official workbook discovery/download failed: " + " | ".join(errors))


def fetch_expected_to_leave_total() -> tuple[list[Point], dict[str, Any]]:
    response, discovery = discover_abs_workbook(
        [
            "https://www.abs.gov.au/statistics/labour/employment-and-unemployment/labour-force-australia/latest-release",
            "https://www.abs.gov.au/statistics/labour/employment-and-unemployment/labour-force-australia-detailed/latest-release",
        ],
        "6291017", "Table 17",
        ["https://www.abs.gov.au/statistics/labour/employment-and-unemployment/labour-force-australia-detailed/mar-2026/6291017.xlsx"],
    )
    (OUT / "abs_6291017_table17_raw.xlsx").write_bytes(response.content)
    points, selected = exact_series_from_workbook(response.content, "A85060262X", response.url, "Q")
    selected.update(discovery)
    selected["note"] = "ABS Table 17 Data1; Does not expect to be with current employer/business in 12 months; Employed total; Persons"
    return points, selected


def fetch_wpi_including_bonuses_yoy(series_id: str = "A2615579C") -> tuple[list[Point], dict[str, Any]]:
    response, discovery = discover_abs_workbook(
        ["https://www.abs.gov.au/statistics/economy/price-indexes-and-inflation/wage-price-index-australia/latest-release#data-downloads"],
        "634507b", "Table 7b",
        [],
    )
    suffix = ".xlsx" if response.content.startswith(b"PK") else ".xls"
    (OUT / f"abs_wpi_table_7b_raw{suffix}").write_bytes(response.content)
    levels, selected = exact_series_from_workbook(response.content, series_id, response.url, "Q")
    level_map = {point.period: point.value for point in levels}
    ordered = sorted(level_map)
    yoy = {ordered[index]: (level_map[ordered[index]] / level_map[ordered[index - 4]] - 1.0) * 100.0 for index in range(4, len(ordered)) if level_map[ordered[index - 4]] != 0}
    if not yoy:
        raise RuntimeError(f"ABS WPI series {series_id} found but YoY could not be calculated")
    points = [Point(period, value, response.url, note=f"Table 7b Data1; series={series_id}; calculated from quarterly index") for period, value in sorted(yoy.items())]
    selected.update(discovery)
    selected["latest_index_levels"] = [{"period": key, "value": level_map[key]} for key in ordered[-12:]]
    selected["note"] = "Official ABS Table 7b; A2615579C Total hourly rates including bonuses, Private and Public, original; YoY from index levels"
    return points, selected


def fetch_westpac_unemployment_expectations() -> tuple[list[Point], dict[str, Any]]:
    url = "https://library.westpaciq.com.au/content/dam/public/westpaciq/secure/economics/documents/aus/2026/07/er20260714BullConsumerSentiment.pdf"
    patterns = [
        r"Unemployment Expectations Index\s+(?:dropped|fell)\s+[0-9]+(?:\.[0-9]+)?%\s+to\s+([0-9]+(?:\.[0-9]+)?)",
        r"unemployment expectations.{0,240}?to\s+([0-9]+(?:\.[0-9]+)?)\s+in July",
    ]
    points = pdf_value(url, patterns, "2026-07", "westpac_consumer_sentiment_july.pdf")
    return points, {"request_url": url, "note": "Official Westpac-MI Consumer Sentiment Bulletin PDF"}


def official_html_value(url: str, patterns: list[str], period: str, raw_name: str) -> list[Point]:
    response = get(url)
    (OUT / raw_name).write_bytes(response.content)
    visible_text = clean(BeautifulSoup(response.text, "html.parser").get_text(" ", strip=True)).replace("−", "-")
    raw_text = html.unescape(response.text).replace("−", "-")
    for haystack in (visible_text, raw_text):
        for pattern in patterns:
            match = re.search(pattern, haystack, re.I | re.S)
            if match:
                value = number(match.group(1))
                if value is not None:
                    return [Point(period, value, response.url)]
    raise RuntimeError("Official HTML value not parsed")



SP_RELEASES_URL = "https://www.pmi.spglobal.com/Public/Release/PressReleases"


def sp_response_to_text(response: requests.Response) -> str:
    content_type = (response.headers.get("content-type") or "").lower()
    if response.content.startswith(b"%PDF") or "application/pdf" in content_type:
        reader = PdfReader(io.BytesIO(response.content))
        return "\n".join(page.extract_text() or "" for page in reader.pages)
    return BeautifulSoup(response.text, "html.parser").get_text("\n", strip=True)


def sp_preceding_context(anchor: Any) -> str:
    parts: list[str] = []
    for element in anchor.previous_elements:
        if isinstance(element, NavigableString):
            text = clean(str(element))
            if text:
                parts.append(text)
        if len(" ".join(parts)) >= 320:
            break
    return " ".join(reversed(parts[-24:]))


def parse_sp_release_date(value: str) -> datetime | None:
    text = clean(value)
    if not text:
        return None
    for format_value in ("%B %d %Y", "%b %d %Y"):
        try:
            return datetime.strptime(text, format_value).replace(tzinfo=timezone.utc)
        except ValueError:
            continue
    return None


def extract_sp_reference_month(text: str) -> str | None:
    head = clean(text[:5000])
    matches = list(re.finditer(
        r"\b(January|February|March|April|May|June|July|August|September|October|November|December)\s+(20\d{2})\b",
        head,
        re.I,
    ))
    if not matches:
        return None
    month_map = {name.lower(): index for index, name in enumerate(
        "January February March April May June July August September October November December".split(), 1
    )}
    match = matches[0]
    return f"{int(match.group(2)):04d}-{month_map[match.group(1).lower()]:02d}"


def discover_australia_pmi_releases() -> list[dict[str, str]]:
    response = get(SP_RELEASES_URL)
    (OUT / "sp_global_release_calendar.html").write_bytes(response.content)
    soup = BeautifulSoup(response.text, "html.parser")
    candidates: list[dict[str, str]] = []
    seen: set[str] = set()
    for anchor in soup.find_all("a", href=True):
        href = anchor.get("href", "")
        if "/Public/Home/PressRelease/" not in href:
            continue
        url = urljoin(SP_RELEASES_URL, href)
        if url in seen:
            continue
        context = clean(" ".join(part for part in (
            anchor.get_text(" ", strip=True),
            anchor.parent.get_text(" ", strip=True) if anchor.parent else "",
            sp_preceding_context(anchor),
        ) if part))
        title_match = re.search(
            r"(S&P Global\s+(?:Flash\s+)?Australia(?:\s+(?:Manufacturing|Services))?\s+PMI(?:[^|]{0,80})?)",
            context,
            re.I,
        )
        if not title_match:
            title_match = re.search(
                r"((?:Flash\s+)?Australia(?:\s+(?:Manufacturing|Services))?\s+PMI(?:[^|]{0,80})?)",
                context,
                re.I,
            )
        if not title_match:
            continue
        title = clean(title_match.group(1))
        release_date_match = re.search(
            r"([A-Za-z]+\s+\d{1,2},?\s+20\d{2})\s+\d{2}:\d{2}\s+UTC",
            context,
            re.I,
        )
        candidates.append({
            "title": title,
            "index_context": context,
            "url": url,
            "release_date": release_date_match.group(1).replace(",", "") if release_date_match else "",
        })
        seen.add(url)
    cutoff = datetime.now(timezone.utc) - timedelta(days=150)
    recent_candidates: list[dict[str, str]] = []
    for candidate in candidates:
        release_date = parse_sp_release_date(candidate.get("release_date", ""))
        if release_date is not None and release_date >= cutoff:
            recent_candidates.append(candidate)
    if not recent_candidates:
        recent_candidates = candidates[:30]
    recent_candidates.sort(
        key=lambda item: parse_sp_release_date(item.get("release_date", ""))
        or datetime.min.replace(tzinfo=timezone.utc),
        reverse=True,
    )
    (OUT / "sp_global_australia_release_candidates.json").write_text(
        json.dumps(recent_candidates, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    log(
        f"[S&P PMI] Australia discovered={len(candidates)} "
        f"recent_to_process={len(recent_candidates)}"
    )
    return recent_candidates


def extract_australia_pmi_value(text: str, sector: str) -> float | None:
    compact = clean(text).replace("™", "").replace("®", "")
    number_pattern = r"([0-9]{1,2}(?:\.[0-9]+)?)"
    if sector == "manufacturing":
        patterns = [
            rf"\bFlash Australia Manufacturing PMI\s*:\s*{number_pattern}\b",
            rf"\bAustralia Manufacturing PMI\s*:\s*{number_pattern}\b",
            r"\bseasonally adjusted S&P Global Australia Manufacturing PMI\s+"
            rf"(?:posted|registered|stood at|rose to|fell to)\s+{number_pattern}\b",
        ]
    else:
        patterns = [
            rf"\bFlash Australia Services PMI Business Activity Index\s*:\s*{number_pattern}\b",
            rf"\bAustralia Services PMI Business Activity Index\s*:\s*{number_pattern}\b",
            r"\bS&P Global Australia Services PMI Business Activity Index\s+"
            rf"(?:posted|registered|stood at|rose to|fell to)\s+{number_pattern}\b",
        ]
    for priority, pattern in enumerate(patterns, 1):
        for match in re.finditer(pattern, compact, re.I):
            value = float(match.group(1))
            context = compact[max(0, match.start() - 120):match.end() + 150]
            if 20.0 <= value <= 80.0 and not (
                value == 50.0 and re.search(r">\s*50|50\s*=", context)
            ):
                log(f"[S&P PMI PARSER] sector={sector} value={value} priority={priority}")
                return value
    return None


def extract_previous_june_value(text: str, sector: str) -> float | None:
    compact = clean(text).replace("™", "").replace("®", "")
    label = (
        r"Flash Australia Manufacturing PMI"
        if sector == "manufacturing"
        else r"Flash Australia Services PMI Business Activity Index"
    )
    match = re.search(
        label + r"\s*:\s*[0-9]{1,2}(?:\.[0-9]+)?\s*\(\s*Jun\s*:\s*([0-9]{1,2}(?:\.[0-9]+)?)\s*\)",
        compact,
        re.I,
    )
    return float(match.group(1)) if match else None


def sp_australia_pmi(label: str) -> tuple[list[Point], dict[str, Any]]:
    sector = "manufacturing" if label == "製造業PMI" else "services"
    candidates = discover_australia_pmi_releases()
    attempts: list[dict[str, Any]] = []
    for candidate in candidates:
        title = candidate.get("title", "").lower()
        if "australia" not in title or "pmi" not in title:
            continue
        try:
            response = get(candidate["url"])
            text = sp_response_to_text(response)
            safe_id = candidate["url"].rstrip("/").split("/")[-1]
            (OUT / f"sp_global_australia_{safe_id}.txt").write_text(text, encoding="utf-8")
            reference_month = extract_sp_reference_month(text)
            current = extract_australia_pmi_value(text, sector)
            previous = extract_previous_june_value(text, sector)
            attempt = {
                "url": response.url,
                "title": candidate.get("title", ""),
                "content_type": response.headers.get("content-type", ""),
                "content_length": len(response.content),
                "text_length": len(text),
                "contains_australia": "Australia" in text,
                "contains_manufacturing": "Manufacturing PMI" in text,
                "contains_services": "Services PMI" in text,
                "reference_month": reference_month,
                "current": current,
                "previous": previous,
                "text_preview": clean(text[:800]),
            }
            attempts.append(attempt)
            if reference_month != "2026-07" or current is None:
                continue
            points = [Point("2026-07", current, response.url, status="flash", note="July 2026 Flash")]
            if previous is not None:
                points.insert(0, Point("2026-06", previous, response.url, status="final", note="Previous value in July Flash release"))
            (OUT / "sp_global_australia_release_parsed.json").write_text(
                json.dumps(attempts, ensure_ascii=False, indent=2), encoding="utf-8"
            )
            return points, {
                "note": "Same calendar discovery and labelled-value parser used by successful UK PMI pipeline",
                "selected_release": attempt,
                "release_candidates": candidates,
            }
        except Exception as error:
            attempts.append({"url": candidate.get("url", ""), "error": f"{type(error).__name__}: {error}"})
    (OUT / "sp_global_australia_release_parsed.json").write_text(
        json.dumps(attempts, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    raise RuntimeError("No Australia PMI value parsed from discovered official S&P releases")

def pdf_value(url: str, patterns: list[str], period: str, raw_name: str) -> list[Point]:
    response = get(url)
    (OUT / raw_name).write_bytes(response.content)
    reader = PdfReader(io.BytesIO(response.content))
    text = clean(" ".join(page.extract_text() or "" for page in reader.pages)).replace("−", "-")
    for pattern in patterns:
        match = re.search(pattern, text, re.I | re.S)
        if match:
            value = number(match.group(1))
            if value is not None:
                return [Point(period, value, response.url)]
    raise RuntimeError("Official PDF value not parsed")


def compare(expected: dict[str, float], points: list[Point], tolerance: float = 0.051) -> dict[str, Any]:
    mapping = {point.period: point for point in points}
    rows = []
    for period, expected_value in sorted(expected.items(), reverse=True):
        point = mapping.get(period)
        difference = None if point is None else point.value - expected_value
        rows.append({
            "period": period,
            "expected": expected_value,
            "official": None if point is None else point.value,
            "difference": difference,
            "match": point is not None and abs(difference) <= tolerance,
        })
    available = [row for row in rows if row["official"] is not None]
    if not expected:
        status = "OFFICIAL_ONLY" if points else "NO_DATA"
    elif len(available) == len(rows) and all(row["match"] for row in rows):
        status = "MATCH_ALL"
    elif available and all(row["match"] for row in available):
        status = "MATCH_AVAILABLE"
    elif available:
        status = "VALUE_MISMATCH"
    else:
        status = "NO_SAME_PERIOD"
    return {"status": status, "rows": rows}


TARGETS = [
    Target("就業新增", "M", {"2026-06": 76.3, "2026-05": 44.0, "2026-04": -38.6}, "API/CSV", "ABS LF"),
    Target("失業率", "M", {"2026-06": 4.428344, "2026-05": 4.3713481, "2026-04": 4.4904846}, "API/CSV", "ABS LF"),
    Target("職缺", "Q", {"2026-05": 329.5}, "API/CSV", "ABS JV"),
    Target("ANZ職缺廣告", "M", {"2026-06": 115.837847, "2026-05": 116.016520, "2026-04": 113.793843}, "XLSX", "ANZ official download"),
    Target("時薪YoY", "Q", {}, "API/CSV", "ABS WPI candidate"),
    Target("預計離職", "A", {}, "XLSX", "ABS Job Mobility"),
    Target("失業預期", "M", {"2026-07": 129.924065, "2026-06": 139.834025, "2026-05": 140.019870, "2026-04": 147.758451}, "HTML", "Westpac-Melbourne Institute official release"),
    Target("CPI YoY", "M", {"2026-06": 3.8, "2026-05": 4.0, "2026-04": 4.2}, "API/CSV", "ABS CPI_MONTHLY"),
    Target("Trimmed Mean YoY", "M", {"2026-06": 3.6, "2026-05": 3.6, "2026-04": 3.4}, "API/CSV", "ABS CPI_MONTHLY"),
    Target("零售", "M", {"2026-05": 1.3, "2026-04": -1.1}, "API/CSV", "ABS HSI_M Monthly Household Spending"),
    Target("NAB企業售價", "M", {"2026-06": 0.59057, "2026-05": 0.92126, "2026-04": 1.79622}, "PDF", "NAB Monthly Business Survey"),
    Target("消費信心", "M", {"2026-07": 83.933514, "2026-06": 80.612096, "2026-05": 82.978906, "2026-04": 80.149167}, "HTML", "Westpac-Melbourne Institute official release"),
    Target("製造業PMI", "M", {"2026-07": 51.7, "2026-06": 51.5, "2026-05": 50.7, "2026-04": 51.3}, "HTML", "S&P Global official releases"),
    Target("服務業PMI", "M", {"2026-07": 53.0, "2026-06": 50.5, "2026-05": 48.7, "2026-04": 50.7}, "HTML", "S&P Global official releases"),
    Target("GDP YoY", "Q", {"2026-Q1": 2.51974}, "API/CSV", "ABS ANA_AGG"),
    Target("GDP私人消費YoY", "Q", {"2026-Q1": 2.4728}, "API/CSV", "ABS ANA_AGG"),
    Target("GDP投資YoY", "Q", {"2026-Q1": 6.47494}, "API/CSV", "ABS ANA_AGG"),
]


def run_target(target: Target) -> tuple[list[Point], dict[str, Any]]:
    label = target.label
    if label == "就業新增":
        return fetch_abs_target("LF", "2025-01", ["employed", "persons", "australia", "seasonally adjusted"], ["rate", "hours", "state"], target.expected, "mom_diff")
    if label == "失業率":
        return fetch_abs_target("LF", "2025-01", ["unemployment rate", "australia", "seasonally adjusted"], ["state", "youth"], target.expected)
    if label == "職缺":
        return fetch_abs_target("JV", "2025-01", ["job vacancies", "australia", "seasonally adjusted", "private and public"], ["trend", "original"], target.expected)
    if label == "ANZ職缺廣告":
        return fetch_anz_job_ads(target.expected)
    if label == "時薪YoY":
        return fetch_wpi_including_bonuses_yoy("A2615579C")
    if label == "預計離職":
        return fetch_expected_to_leave_total()
    if label == "CPI YoY":
        return fetch_abs_target("CPI_MONTHLY", "2025-01", ["all groups", "australia", "annual"], ["trimmed", "quarterly"], target.expected)
    if label == "Trimmed Mean YoY":
        return fetch_abs_target("CPI_MONTHLY", "2025-01", ["trimmed mean", "australia", "annual"], [], target.expected)
    if label == "零售":
        return fetch_abs_target("HSI_M", "2025-01", ["household spending", "australia", "seasonally adjusted", "monthly"], ["trend", "through the year"], target.expected)
    if label == "NAB企業售價":
        releases = [
            ("2026-06", "https://www.nab.com.au/content/dam/nab/documents/news/2026m06-nab-monthly-business-survey-mnb.pdf", "nab_june_2026.pdf"),
            ("2026-05", "https://news.nab.com.au/content/dam/nab-news/documents/economics/nab-monthly-business-survey-may-26.pdf", "nab_may_2026.pdf"),
            ("2026-04", "https://news.nab.com.au/content/dam/nab-news/documents/economics/202604%20NAB%20Monthly%20Business%20Survey%20April.pdf", "nab_april_2026.pdf"),
        ]
        points = []
        errors = []
        patterns = [r"Product prices\s+To\s*([+-]?\d+(?:\.\d+)?)%", r"product price growth.{0,120}?(?:at|to)\s*([+-]?\d+(?:\.\d+)?)%", r"final product prices.{0,100}?([+-]?\d+(?:\.\d+)?)%"]
        for period, url, raw_name in releases:
            try:
                points.extend(pdf_value(url, patterns, period, raw_name))
            except Exception as error:
                errors.append(f"{period}: {error}")
        if not points:
            raise RuntimeError("NAB monthly PDFs failed: " + " | ".join(errors))
        return points, {"note": "Official monthly PDFs; no public API/CSV found", "errors": errors}
    if label == "消費信心":
        releases = [
            (
                "2026-07",
                "https://www.westpac.com.au/news/making-news/2026/07/feeling-better-but-not-better-off-consumer-sentiment-and-the-cost-of-living-crunch/",
                [r"Consumer Sentiment Index rose.{0,120}?to\s*([0-9]+(?:\.\d+)?)"],
                "westpac_consumer_sentiment_july.html",
            ),
            (
                "2026-06",
                "https://www.westpac.com.au/news/making-news/2026/06/consumer-sentiment-slips-again-as-cost-of-living-pressures-weigh-on-households/",
                [r"Consumer Sentiment Index dropped.{0,120}?to\s*([0-9]+(?:\.\d+)?)"],
                "westpac_consumer_sentiment_june.html",
            ),
        ]
        points = []
        errors = []
        for period, url, patterns, raw_name in releases:
            try:
                points.extend(official_html_value(url, patterns, period, raw_name))
            except Exception as error:
                errors.append(f"{period}: {error}")
        if not points:
            raise RuntimeError("Westpac official releases failed: " + " | ".join(errors))
        return points, {"note": "Official Westpac releases; full precision history is licensed", "errors": errors}
    if label == "失業預期":
        return fetch_westpac_unemployment_expectations()
    if label in {"製造業PMI", "服務業PMI"}:
        return sp_australia_pmi(label)
    if label == "GDP YoY":
        return fetch_abs_target("ANA_AGG", "2024-Q1", ["gross domestic product", "chain volume", "index", "seasonally adjusted"], ["per capita", "percentage changes"], target.expected, "yoy_pct_q")
    if label == "GDP私人消費YoY":
        return fetch_abs_workbook_candidate(
            ["https://www.abs.gov.au/statistics/economy/national-accounts/australian-national-accounts-national-income-expenditure-and-product/mar-2026/5206008_Household_Final_Consumption_Expenditure.xlsx"],
            target.expected,
            "abs_5206008_household_consumption.xlsx",
            "ABS Table 8 Household Final Consumption Expenditure official XLSX",
        )
    if label == "GDP投資YoY":
        return fetch_abs_workbook_candidate(
            [
                "https://www.abs.gov.au/statistics/economy/national-accounts/australian-national-accounts-national-income-expenditure-and-product/mar-2026/5206002_expenditure_volume_measures.xlsx",
                "https://www.abs.gov.au/statistics/economy/national-accounts/australian-national-accounts-national-income-expenditure-and-product/mar-2026/5206002_expenditure_volume_measures.xls",
            ],
            target.expected,
            "abs_5206002_expenditure_volume_measures.xlsx",
            "ABS Table 2 Expenditure on GDP, chain volume measures official workbook",
        )
    raise RuntimeError(f"No test mapping for {label}")


def main() -> int:
    OUT.mkdir(parents=True, exist_ok=True)
    report: dict[str, Any] = {
        "script_version": VERSION,
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "source_policy": "official API/CSV/XLSX > official PDF > official HTML",
        "results": [],
        "summary": {},
    }
    for target in TARGETS:
        log(f"\n[TEST] {target.label} | {target.source_name}")
        item: dict[str, Any] = {"target": asdict(target)}
        try:
            points, diagnostics = run_target(target)
            points = dedupe(points)
            comparison = compare(target.expected, points)
            item.update({
                "status": comparison["status"],
                "comparison": comparison,
                "latest_points": [asdict(point) for point in points[-12:]],
                "diagnostics": diagnostics,
            })
            log(f"[RESULT] {item['status']} points={len(points)}")
        except Exception as error:
            item.update({
                "status": "FETCH_ERROR",
                "error": f"{type(error).__name__}: {error}",
                "traceback": traceback.format_exc(limit=10),
            })
            log(f"[ERROR] {item['error']}")
        report["summary"][item["status"]] = report["summary"].get(item["status"], 0) + 1
        report["results"].append(item)

    (OUT / "au_source_comparison.json").write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    with (OUT / "au_source_comparison.csv").open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(["indicator", "status", "source_tier", "source", "period", "AU_ECON", "official", "difference", "match", "error"])
        for item in report["results"]:
            target = item["target"]
            rows = item.get("comparison", {}).get("rows", [])
            if not rows:
                writer.writerow([target["label"], item["status"], target["source_tier"], target["source_name"], "", "", "", "", "", item.get("error", "")])
            for row in rows:
                writer.writerow([target["label"], item["status"], target["source_tier"], target["source_name"], row["period"], row["expected"], row["official"], row["difference"], row["match"], item.get("error", "")])

    log("\n=== SUMMARY ===")
    for status, count in sorted(report["summary"].items()):
        log(f"{status}: {count}")
    log(f"Output: {OUT}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
