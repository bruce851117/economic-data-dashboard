#!/usr/bin/env python3
"""Targeted validation of the remaining EU structured macro sources.

Tests only the six sources that still need a machine-readable implementation:
1. Spain core CPI: INE table 50907 JSON
2. Spain unemployment rate: INE table 65219/65334 JSON
3. France core CPI: INSEE BDM idbank 011814143
4. Germany core CPI: Destatis GENESIS REST table 61111-0006
5. Germany industrial production: Destatis GENESIS REST table 42153-0001
6. Germany ifo business climate: official XLS/XLSX, official PDF fallback

The script never writes production data. It produces diagnostics and compares only
periods already listed in EXPECTED. HTML is used only to discover official download
assets; values are read from JSON/XML/CSV/XLS/XLSX/PDF.
"""
from __future__ import annotations

import argparse
import csv
import io
import json
import math
import os
import re
import sys
import time
import traceback
import unicodedata
import xml.etree.ElementTree as ET
from dataclasses import asdict, dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Callable
from urllib.parse import urljoin

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

VERSION = "2026-07-30-remaining-structured-v1"
TIMEOUT = 60
UA = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/126 Safari/537.36"

EXPECTED: dict[str, dict[str, float]] = {
    "西Core CPI": {"2026-06": 2.9, "2026-05": 3.0, "2026-04": 2.8},
    "西 失業率": {"2026-Q2": 9.87, "2026-Q1": 10.83},
    "法 Core CPI": {"2026-06": 1.00498, "2026-05": 1.25824, "2026-04": 1.15814},
    "德 Core CPI": {"2026-06": 2.45139, "2026-05": 2.54022, "2026-04": 2.29007},
    "德 工業": {"2026-05": 0.0, "2026-04": -0.8762322015334},
    "德 企業信心": {"2026-07": 86.59582, "2026-06": 85.7, "2026-05": 85.0, "2026-04": 84.5},
}

HTTP_LOG: list[dict[str, Any]] = []
DETAILS: dict[str, Any] = {}


def clean(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def norm(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", " ", unicodedata.normalize("NFKD", clean(value)).encode("ascii", "ignore").decode().lower()).strip()


def num(value: Any) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value) if math.isfinite(float(value)) else None
    text = clean(value).replace("−", "-").replace("%", "")
    if not text or text.lower() in {"nan", "null", "none", "..", "...", ":", "-"}:
        return None
    text = re.sub(r"[^0-9,.-]", "", text)
    if "," in text and "." not in text:
        text = text.replace(",", ".")
    elif "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".") if text.rfind(",") > text.rfind(".") else text.replace(",", "")
    try:
        return float(text)
    except ValueError:
        return None


def period_key(value: Any, year: Any = None, period_code: Any = None) -> str | None:
    text = clean(value)
    if re.fullmatch(r"1\d{12}", text):
        return datetime.fromtimestamp(int(text) / 1000, timezone.utc).strftime("%Y-%m")
    q = re.search(r"(20\d{2})\D*[QqTt]([1-4])", text)
    if q:
        return f"{q.group(1)}-Q{q.group(2)}"
    m = re.search(r"(20\d{2})[-/. ](0?[1-9]|1[0-2])(?:\D|$)", text)
    if m:
        return f"{m.group(1)}-{int(m.group(2)):02d}"
    if year:
        try:
            y = int(year)
            p = int(period_code)
            if 1 <= p <= 12:
                return f"{y:04d}-{p:02d}"
            if 19 <= p <= 22:
                return f"{y:04d}-Q{p - 18}"
        except (TypeError, ValueError):
            pass
    return None


@dataclass
class Point:
    period: str
    value: float
    source_url: str
    note: str = ""


def dedupe(points: list[Point]) -> list[Point]:
    return [p for _, p in sorted({p.period: p for p in points}.items())]


def make_session() -> requests.Session:
    s = requests.Session()
    retry = Retry(total=2, connect=2, read=2, backoff_factor=1.0,
                  status_forcelist=(429, 500, 502, 503, 504), allowed_methods=("GET", "POST"))
    s.mount("https://", HTTPAdapter(max_retries=retry))
    s.headers.update({"User-Agent": UA, "Accept": "*/*"})
    return s


HTTP = make_session()


def request(method: str, url: str, **kwargs: Any) -> requests.Response:
    started = time.time()
    try:
        r = HTTP.request(method, url, timeout=TIMEOUT, **kwargs)
        HTTP_LOG.append({"method": method, "url": r.url, "status": r.status_code,
                         "bytes": len(r.content), "seconds": round(time.time() - started, 3),
                         "content_type": r.headers.get("content-type", "")})
        r.raise_for_status()
        return r
    except Exception as exc:
        HTTP_LOG.append({"method": method, "url": url, "status": None,
                         "seconds": round(time.time() - started, 3), "error": f"{type(exc).__name__}: {exc}"})
        raise


# ------------------------- INE table API -------------------------

def ine_table(table_id: str) -> tuple[list[dict[str, Any]], str]:
    url = f"https://servicios.ine.es/wstempus/js/ES/DATOS_TABLA/{table_id}"
    errors = []
    for params in ({"tip": "AM", "nult": 24}, {"nult": 24}, {}):
        try:
            r = request("GET", url, params=params)
            payload = r.json()
            if isinstance(payload, list) and payload:
                return payload, r.url
        except Exception as exc:
            errors.append(f"{params}: {exc}")
    raise RuntimeError("INE DATOS_TABLA failed: " + " | ".join(errors))


def ine_series_points(series: dict[str, Any], source_url: str) -> list[Point]:
    rows = series.get("Data") or series.get("data") or []
    points: list[Point] = []
    for row in rows:
        period = period_key(row.get("Fecha"), row.get("Anyo"), row.get("FK_Periodo"))
        value = num(row.get("Valor"))
        if period and value is not None:
            points.append(Point(period, value, source_url))
    return dedupe(points)


def choose_ine_series(table_id: str, label: str, required: list[str], excluded: list[str] | None = None) -> list[Point]:
    payload, source_url = ine_table(table_id)
    excluded = excluded or []
    candidates = []
    for series in payload:
        name = clean(series.get("Nombre") or series.get("name"))
        code = clean(series.get("COD") or series.get("Codigo") or series.get("Id"))
        n = norm(name)
        if not all(term in n for term in required) or any(term in n for term in excluded):
            continue
        points = ine_series_points(series, source_url)
        mapping = {p.period: p.value for p in points}
        shared = [period for period in EXPECTED[label] if period in mapping]
        error = sum(abs(mapping[p] - EXPECTED[label][p]) for p in shared) / len(shared) if shared else None
        candidates.append({"code": code, "name": name, "shared": shared,
                           "mean_abs_error": error, "points": [asdict(p) for p in points[-30:]]})
    candidates.sort(key=lambda x: (-len(x["shared"]), x["mean_abs_error"] if x["mean_abs_error"] is not None else 1e9))
    DETAILS[f"INE table {table_id} candidates"] = candidates[:50]
    if not candidates or not candidates[0]["shared"]:
        raise RuntimeError(f"INE table {table_id}: target series/period not found")
    best = candidates[0]
    return [Point(x["period"], x["value"], source_url,
                  f"INE table={table_id}; series={best['code']}; {best['name']}") for x in best["points"]]


def spain_core_cpi() -> list[Point]:
    return choose_ine_series("50907", "西Core CPI",
                             ["general sin alimentos no elaborados ni productos energeticos", "variacion anual"])


def spain_unemployment() -> list[Point]:
    errors = []
    configurations = [
        ("65219", ["ambos sexos", "total"], ["hombres", "mujeres"]),
        ("65334", ["ambos sexos", "total nacional", "total"], ["hombres", "mujeres"]),
    ]
    for table_id, required, excluded in configurations:
        try:
            return choose_ine_series(table_id, "西 失業率", required, excluded)
        except Exception as exc:
            errors.append(f"{table_id}: {exc}")
    raise RuntimeError("Spanish unemployment tables failed: " + " | ".join(errors))


# ------------------------- INSEE BDM -------------------------

def insee_idbank(idbank: str) -> list[Point]:
    url = f"https://api.insee.fr/series/BDM/V1/data/SERIES_BDM/{idbank}"
    r = request("GET", url, params={"lastNObservations": 30})
    root = ET.fromstring(r.content)
    points = []
    for obs in root.iter():
        if obs.tag.endswith("Obs"):
            period = period_key(obs.attrib.get("TIME_PERIOD") or obs.attrib.get("timePeriod"))
            value = num(obs.attrib.get("OBS_VALUE") or obs.attrib.get("obsValue"))
            if period and value is not None:
                points.append(Point(period, value, r.url, f"INSEE idbank={idbank}; level"))
    if not points:
        raise RuntimeError(f"INSEE idbank {idbank}: no observations")
    return dedupe(points)


def yoy_from_levels(points: list[Point], expected_periods: list[str], note: str) -> list[Point]:
    levels = {p.period: p for p in points}
    out = []
    for period in expected_periods:
        if "-Q" in period:
            year, qtr = period.split("-Q")
            prior = f"{int(year)-1:04d}-Q{qtr}"
        else:
            year, month = period.split("-")
            prior = f"{int(year)-1:04d}-{month}"
        if period in levels and prior in levels and levels[prior].value:
            value = (levels[period].value / levels[prior].value - 1) * 100
            out.append(Point(period, value, levels[period].source_url, note))
    if not out:
        raise RuntimeError("Cannot calculate YoY from official levels")
    return out


def france_core_cpi() -> list[Point]:
    levels = insee_idbank("011814143")
    DETAILS["INSEE 011814143 levels"] = [asdict(p) for p in levels]
    return yoy_from_levels(levels, list(EXPECTED["法 Core CPI"]),
                           "YoY calculated from INSEE Base 2025 underlying CPI levels; idbank=011814143")


# ------------------------- Destatis GENESIS REST -------------------------

def decode_csv(content: bytes) -> list[list[str]]:
    text = None
    for encoding in ("utf-8-sig", "cp1252", "iso-8859-1"):
        try:
            candidate = content.decode(encoding)
            if "\ufffd" not in candidate:
                text = candidate
                break
        except UnicodeDecodeError:
            pass
    text = text or content.decode("cp1252", errors="replace")
    delimiter = ";" if text[:10000].count(";") > text[:10000].count(",") else ","
    return list(csv.reader(io.StringIO(text), delimiter=delimiter))


def genesis_credentials() -> dict[str, str]:
    token = clean(os.getenv("GENESIS_TOKEN"))
    username = clean(os.getenv("GENESIS_USERNAME"))
    password = clean(os.getenv("GENESIS_PASSWORD"))
    if token:
        return {"username": token, "password": ""}
    if username:
        return {"username": username, "password": password}
    return {"username": "GUEST", "password": ""}


def genesis_table(table_id: str, startyear: int = 2025, endyear: int = 2026) -> tuple[list[list[str]], str]:
    """Request a full GENESIS table through the official REST interface.

    GENESIS_TOKEN is preferred. GENESIS_USERNAME and GENESIS_PASSWORD are also
    supported. Guest mode is attempted so the diagnostic remains executable.
    """
    endpoints = [
        "https://www-genesis.destatis.de/genesisWS/rest/2020/data/tablefile",
        "https://genesis.destatis.de/genesisWS/rest/2020/data/tablefile",
    ]
    auth = genesis_credentials()
    body = {**auth, "name": table_id, "area": "all", "compress": "false",
            "startyear": str(startyear), "endyear": str(endyear),
            "timeslices": "", "regionalvariable": "", "regionalkey": "",
            "classifyingvariable1": "", "classifyingkey1": "",
            "classifyingvariable2": "", "classifyingkey2": "",
            "classifyingvariable3": "", "classifyingkey3": "",
            "format": "ffcsv", "job": "false", "language": "de"}
    errors = []
    for endpoint in endpoints:
        try:
            r = request("POST", endpoint, json=body)
            content_type = r.headers.get("content-type", "").lower()
            if "json" in content_type or r.text.lstrip().startswith("{"):
                payload = r.json()
                DETAILS[f"GENESIS response {table_id}"] = payload
                obj = payload.get("Object") or payload.get("object") or {}
                content = obj.get("Content") or obj.get("content")
                if isinstance(content, str) and content.strip():
                    return list(csv.reader(io.StringIO(content), delimiter=";")), r.url
                raise RuntimeError(clean(payload.get("Status") or payload.get("status") or payload)[:500])
            return decode_csv(r.content), r.url
        except Exception as exc:
            errors.append(f"{endpoint}: {type(exc).__name__}: {exc}")
    # Public flat-file fallback is retained for diagnostics, though it may contain
    # only the site's default table selection.
    fallback = f"https://genesis.destatis.de/genesisWS/downloads/00/tables/{table_id}_00.csv"
    try:
        r = request("GET", fallback)
        rows = decode_csv(r.content)
        DETAILS[f"GENESIS fallback {table_id}"] = rows[:300]
        return rows, r.url
    except Exception as exc:
        errors.append(f"{fallback}: {exc}")
    raise RuntimeError("GENESIS table request failed: " + " | ".join(errors))


def row_periods(rows: list[list[str]]) -> dict[int, str]:
    years: dict[int, int] = {}
    periods: dict[int, str] = {}
    width = max((len(r) for r in rows[:30]), default=0)
    months = {name: i for i, name in enumerate(
        "januar februar marz april mai juni juli august september oktober november dezember".split(), 1)}
    for row in rows[:30]:
        for col, cell in enumerate(row):
            text = norm(cell)
            if re.fullmatch(r"20\d{2}", text):
                years[col] = int(text)
    current = None
    for col in range(width):
        if col in years:
            current = years[col]
        elif current is not None:
            years[col] = current
    for row in rows[:30]:
        for col, cell in enumerate(row):
            text = norm(cell)
            year = years.get(col)
            if year and text in months:
                periods[col] = f"{year:04d}-{months[text]:02d}"
    return periods


def select_destatis_row(rows: list[list[str]], label: str, required: list[str], excluded: list[str] | None = None) -> list[Point]:
    excluded = excluded or []
    periods = row_periods(rows)
    candidates = []
    for idx, row in enumerate(rows):
        text = norm(" ".join(clean(x) for x in row[:12]))
        if not all(term in text for term in required) or any(term in text for term in excluded):
            continue
        levels = {period: num(row[col]) for col, period in periods.items()
                  if col < len(row) and num(row[col]) is not None}
        if levels:
            candidates.append({"row": idx + 1, "label": " | ".join(clean(x) for x in row[:8]), "levels": levels})
    DETAILS[f"Destatis {label} candidates"] = candidates[:50]
    if not candidates:
        raise RuntimeError(f"Destatis {label}: target row missing; GENESIS_TOKEN may be required for full table")
    best = max(candidates, key=lambda x: len(x["levels"]))
    source = "official GENESIS REST/CSV"
    points = [Point(period, value, source, f"row={best['row']}; {best['label']}") for period, value in best["levels"].items()]
    return dedupe(points)


def germany_core_cpi() -> list[Point]:
    rows, source_url = genesis_table("61111-0006", 2025, 2026)
    DETAILS["Destatis 61111-0006 raw preview"] = rows[:300]
    levels = select_destatis_row(rows, "core CPI",
                                 ["ohne", "nahrungsmittel", "energie"],
                                 ["alkohol", "tabak"])
    for p in levels:
        p.source_url = source_url
    return yoy_from_levels(levels, list(EXPECTED["德 Core CPI"]),
                           "YoY from GENESIS CPI special-aggregate levels")


def germany_industry() -> list[Point]:
    rows, source_url = genesis_table("42153-0001", 2025, 2026)
    DETAILS["Destatis 42153-0001 raw preview"] = rows[:300]
    levels = select_destatis_row(rows, "industry",
                                 ["produzierendes gewerbe", "kalenderbereinigt"],
                                 ["saisonbereinigt"])
    for p in levels:
        p.source_url = source_url
    return yoy_from_levels(levels, list(EXPECTED["德 工業"]),
                           "YoY from GENESIS calendar-adjusted production levels")


# ------------------------- ifo XLS/XLSX and PDF fallback -------------------------

def workbook_rows(content: bytes, url: str) -> list[dict[str, Any]]:
    if url.lower().split("?")[0].endswith(".xls") or content[:8] == bytes.fromhex("D0CF11E0A1B11AE1"):
        import xlrd
        book = xlrd.open_workbook(file_contents=content)
        return [{"sheet": sh.name, "rows": [sh.row_values(i) for i in range(sh.nrows)]} for sh in book.sheets()]
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    return [{"sheet": ws.title, "rows": [list(r) for r in ws.iter_rows(values_only=True)]} for ws in wb.worksheets]


def asset_urls(page_url: str) -> list[str]:
    r = request("GET", page_url)
    soup = BeautifulSoup(r.text, "html.parser")
    urls: list[str] = []
    for tag in soup.find_all(True):
        for attr, value in tag.attrs.items():
            values = value if isinstance(value, list) else [value]
            for raw in values:
                text = clean(raw)
                if not text:
                    continue
                for match in re.findall(r"(?:https?:)?//[^\"'\s<>]+|/[^\"'\s<>]+", text):
                    candidate = urljoin(r.url, match)
                    if candidate.lower().split("?")[0].endswith((".xlsx", ".xls", ".pdf")) or "/media/" in candidate or "/download" in candidate:
                        urls.append(candidate)
    # Raw HTML regex captures escaped JSON asset URLs.
    raw_html = r.text.replace("\\/", "/")
    for match in re.findall(r"(?:https?:)?//[^\"'\s<>]+|/[^\"'\s<>]+", raw_html):
        candidate = urljoin(r.url, match)
        if candidate.lower().split("?")[0].endswith((".xlsx", ".xls", ".pdf")) or "/media/" in candidate:
            urls.append(candidate)
    return list(dict.fromkeys(urls))


def ifo_from_workbook(content: bytes, url: str) -> list[Point]:
    books = workbook_rows(content, url)
    DETAILS[f"ifo workbook preview {url}"] = [{"sheet": b["sheet"], "rows": b["rows"][:30]} for b in books]
    candidates = []
    for book in books:
        for idx, row in enumerate(book["rows"]):
            if not row:
                continue
            period = period_key(row[0])
            if not period:
                continue
            for col, cell in enumerate(row[1:], 1):
                value = num(cell)
                if value is None:
                    continue
                context_rows = book["rows"][max(0, idx - 5):idx + 1]
                context = norm(f"{book['sheet']} " + " ".join(clean(x) for rr in context_rows for x in rr[:col + 1]))
                score = sum(3 for term in ["deutschland", "geschaftsklima", "klima", "index", "saisonbereinigt"] if term in context)
                if 70 <= value <= 110:
                    candidates.append({"period": period, "value": value, "score": score,
                                       "sheet": book["sheet"], "row": idx + 1, "col": col + 1})
    DETAILS[f"ifo workbook candidates {url}"] = sorted(candidates, key=lambda x: x["score"], reverse=True)[:100]
    points = []
    for period in EXPECTED["德 企業信心"]:
        rows = [x for x in candidates if x["period"] == period]
        if rows:
            best = max(rows, key=lambda x: x["score"])
            points.append(Point(period, best["value"], url,
                                f"sheet={best['sheet']}; row={best['row']}; col={best['col']}"))
    if not points:
        raise RuntimeError("ifo workbook parsed but target periods not found")
    return dedupe(points)


def ifo_from_pdf(content: bytes, url: str) -> list[Point]:
    from pypdf import PdfReader
    reader = PdfReader(io.BytesIO(content))
    text = "\n".join(page.extract_text() or "" for page in reader.pages)
    DETAILS[f"ifo PDF text {url}"] = text[:50000]
    points = []
    # Press-release sentence is reliable for the latest month; the embedded table
    # remains in DETAILS for deterministic parser refinement if needed.
    months = {name: i for i, name in enumerate(
        "januar februar marz april mai juni juli august september oktober november dezember".split(), 1)}
    pattern = r"geschaftsklimaindex\s+(?:stieg|sank).*?im\s+(januar|februar|marz|april|mai|juni|juli|august|september|oktober|november|dezember).*?auf\s+([0-9]+[,.][0-9]+)\s+punkte"
    match = re.search(pattern, norm(text), re.S)
    if match:
        month = months[match.group(1)]
        year_match = re.search(r"\b(20\d{2})\b", text)
        if year_match:
            points.append(Point(f"{year_match.group(1)}-{month:02d}", float(match.group(2).replace(",", ".")), url,
                                "ifo official press-release PDF sentence"))
    if not points:
        raise RuntimeError("ifo PDF downloaded but latest value not parsed")
    return points


def ifo_business() -> list[Point]:
    pages = [
        "https://www.ifo.de/ifo-zeitreihen",
        "https://www.ifo.de/umfragen/zeitreihen",
        "https://www.ifo.de/fakten/2026-07-27/ifo-geschaeftsklimaindex-gestiegen-juli-2026",
    ]
    urls = []
    errors = []
    for page in pages:
        try:
            urls.extend(asset_urls(page))
        except Exception as exc:
            errors.append(f"{page}: {exc}")
    urls = list(dict.fromkeys(urls))
    DETAILS["ifo discovered assets"] = urls
    # Workbook assets first.
    for url in urls:
        if not url.lower().split("?")[0].endswith((".xlsx", ".xls")):
            continue
        try:
            r = request("GET", url)
            return ifo_from_workbook(r.content, r.url)
        except Exception as exc:
            errors.append(f"{url}: {exc}")
    # Official PDF/media fallback.
    for url in urls:
        try:
            r = request("GET", url)
            if "pdf" in r.headers.get("content-type", "").lower() or r.content.startswith(b"%PDF"):
                return ifo_from_pdf(r.content, r.url)
        except Exception as exc:
            errors.append(f"{url}: {exc}")
    raise RuntimeError("ifo official asset not parsed: " + " | ".join(errors[:15]))


# ------------------------- reporting -------------------------
TESTS: list[tuple[str, str, Callable[[], list[Point]]]] = [
    ("西Core CPI", "INE table 50907 JSON", spain_core_cpi),
    ("西 失業率", "INE table 65219/65334 JSON", spain_unemployment),
    ("法 Core CPI", "INSEE BDM 011814143 XML", france_core_cpi),
    ("德 Core CPI", "Destatis GENESIS 61111-0006 REST/CSV", germany_core_cpi),
    ("德 工業", "Destatis GENESIS 42153-0001 REST/CSV", germany_industry),
    ("德 企業信心", "ifo official XLS/XLSX; PDF fallback", ifo_business),
]


def compare(label: str, points: list[Point], tolerance: float) -> dict[str, Any]:
    mapping = {p.period: p for p in points}
    rows = []
    for period, expected in sorted(EXPECTED[label].items(), reverse=True):
        point = mapping.get(period)
        if point is None:
            rows.append({"period": period, "EU_ECON": expected, "official": None,
                         "difference": None, "match": False, "note": "official period absent"})
        else:
            difference = point.value - expected
            rows.append({"period": period, "EU_ECON": expected, "official": point.value,
                         "difference": difference, "match": abs(difference) <= tolerance,
                         "source_url": point.source_url, "note": point.note})
    available = [r for r in rows if r["official"] is not None]
    if len(available) == len(rows) and all(r["match"] for r in available):
        status = "MATCH_ALL"
    elif available and all(r["match"] for r in available):
        status = "MATCH_AVAILABLE_PERIODS"
    elif available:
        status = "VALUE_MISMATCH"
    else:
        status = "NO_SAME_PERIOD"
    return {"status": status, "rows": rows}


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--out", default="debug/eu_remaining_structured_sources")
    parser.add_argument("--tolerance", type=float, default=0.051)
    args = parser.parse_args()
    out = Path(args.out)
    out.mkdir(parents=True, exist_ok=True)
    report = {"script_version": VERSION, "generated_at": datetime.now(timezone.utc).isoformat(),
              "genesis_auth_mode": "token" if os.getenv("GENESIS_TOKEN") else ("username" if os.getenv("GENESIS_USERNAME") else "guest"),
              "results": [], "summary": {}}
    for label, source_name, fetcher in TESTS:
        print(f"\n[TEST] {label} | {source_name}", flush=True)
        item = {"label": label, "source_name": source_name}
        try:
            points = dedupe(fetcher())
            item["points"] = [asdict(p) for p in points]
            item["comparison"] = compare(label, points, args.tolerance)
            item["status"] = item["comparison"]["status"]
            print(f"[RESULT] {item['status']} | points={len(points)}", flush=True)
        except Exception as exc:
            item["status"] = "FETCH_ERROR"
            item["error"] = f"{type(exc).__name__}: {exc}"
            item["traceback"] = traceback.format_exc(limit=10)
            print(f"[ERROR] {item['error']}", flush=True)
        report["summary"][item["status"]] = report["summary"].get(item["status"], 0) + 1
        report["results"].append(item)

    (out / "remaining_source_comparison.json").write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    (out / "remaining_source_details.json").write_text(json.dumps(DETAILS, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    (out / "remaining_source_http_log.json").write_text(json.dumps(HTTP_LOG, ensure_ascii=False, indent=2), encoding="utf-8")
    with (out / "remaining_source_comparison.csv").open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["label", "status", "source", "period", "EU_ECON", "official", "difference", "match", "source_url", "note_or_error"])
        for item in report["results"]:
            if "comparison" not in item:
                writer.writerow([item["label"], item["status"], item["source_name"], "", "", "", "", "", "", item.get("error", "")])
                continue
            for row in item["comparison"]["rows"]:
                writer.writerow([item["label"], item["status"], item["source_name"], row["period"], row["EU_ECON"],
                                 row["official"], row["difference"], row["match"], row.get("source_url", ""), row.get("note", "")])
    print("\n=== SUMMARY ===", flush=True)
    for status, count in report["summary"].items():
        print(f"{status}: {count}", flush=True)
    return 0


if __name__ == "__main__":
    sys.exit(main())
